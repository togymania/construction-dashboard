"""Expense CRUD endpoints + Excel bulk import."""
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentUser, DBSession, require_roles
from app.models.budget import BudgetItem
from app.models.budget_category import BudgetCategory
from app.models.expense import Expense, ExpenseStatus
from app.models.project import Project
from app.models.user import User, UserRole
from app.schemas.expense import (
    ExpenseCreate,
    ExpenseImportResult,
    ExpenseResponse,
    ExpenseUpdate,
    ImportRowError,
)

router = APIRouter(tags=["Expenses"])

MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024  # 5 MB

# ---- Column-name mapping (case-insensitive, supports TR & EN) ----

_COLUMN_ALIASES: dict[str, list[str]] = {
    "vendor": ["vendor", "company", "şirket", "şirket adı", "firma", "firma adı"],
    "invoice_number": [
        "invoice",
        "invoice_number",
        "invoice no",
        "invoice #",
        "fatura",
        "fatura no",
        "fatura numarası",
    ],
    "expense_date": [
        "date",
        "payment date",
        "expense_date",
        "tarih",
        "ödeme tarihi",
        "ödeme günü",
    ],
    "amount": ["amount", "total", "tutar", "miktar", "fatura tutarı"],
    "category": [
        "category",
        "kategori",
        "budget category",
        "bütçe kategorisi",
    ],
    "description": ["description", "açıklama", "desc", "note", "not"],
}


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Map canonical field names to column indices using fuzzy aliases."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        normalised = str(raw).strip().lower()
        for field, aliases in _COLUMN_ALIASES.items():
            if normalised in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


# ---- Helpers ----


async def _ensure_project(db, project_id: int) -> Project:
    project = await db.get(Project, project_id)
    if project is None or not project.is_active:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Project not found")
    return project


async def _load_expense(db, expense_id: int) -> Expense:
    stmt = (
        select(Expense)
        .options(
            selectinload(Expense.category),
            selectinload(Expense.creator),
        )
        .where(Expense.id == expense_id)
    )
    result = await db.execute(stmt)
    expense = result.scalar_one_or_none()
    if expense is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")
    return expense


def _expense_to_response(exp: Expense) -> ExpenseResponse:
    return ExpenseResponse(
        id=exp.id,
        project_id=exp.project_id,
        budget_item_id=exp.budget_item_id,
        category_id=exp.category_id,
        category={
            "id": exp.category.id,
            "name": exp.category.name,
            "slug": exp.category.slug,
        },
        description=exp.description,
        amount=exp.amount,
        expense_date=exp.expense_date,
        vendor=exp.vendor,
        invoice_number=exp.invoice_number,
        notes=exp.notes,
        status=exp.status.value,
        creator_name=exp.creator.full_name,
        created_at=exp.created_at,
        updated_at=exp.updated_at,
    )


async def _get_category_by_name(db, name: str) -> BudgetCategory | None:
    """Look up a category by name (case-insensitive)."""
    stmt = select(BudgetCategory).where(
        func.lower(BudgetCategory.name) == name.strip().lower(),
        BudgetCategory.is_active == True,  # noqa: E712
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


async def _get_category_map(db) -> dict[str, int]:
    """Return {lowercase_name: id} for all active categories."""
    stmt = select(BudgetCategory).where(BudgetCategory.is_active == True)  # noqa: E712
    rows = (await db.execute(stmt)).scalars().all()
    return {cat.name.lower(): cat.id for cat in rows}


# ---- Endpoints ----


@router.get(
    "/projects/{project_id}/expenses",
    response_model=list[ExpenseResponse],
    summary="List expenses for a project (with optional filters)",
)
async def list_expenses(
    project_id: int,
    user: CurrentUser,
    db: DBSession,
    category_id: int | None = Query(None),
    date_from: date | None = Query(None),
    date_to: date | None = Query(None),
    search: str | None = Query(None, max_length=200),
) -> list[ExpenseResponse]:
    await _ensure_project(db, project_id)

    stmt = (
        select(Expense)
        .options(
            selectinload(Expense.category),
            selectinload(Expense.creator),
        )
        .where(Expense.project_id == project_id)
        .order_by(Expense.expense_date.desc(), Expense.id.desc())
    )

    if category_id is not None:
        stmt = stmt.where(Expense.category_id == category_id)
    if date_from is not None:
        stmt = stmt.where(Expense.expense_date >= date_from)
    if date_to is not None:
        stmt = stmt.where(Expense.expense_date <= date_to)
    if search:
        pattern = f"%{search}%"
        stmt = stmt.where(
            Expense.description.ilike(pattern) | Expense.vendor.ilike(pattern)
        )

    result = await db.execute(stmt)
    return [_expense_to_response(e) for e in result.scalars().all()]


@router.post(
    "/projects/{project_id}/expenses",
    response_model=ExpenseResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Create a single expense (always PAID)",
)
async def create_expense(
    project_id: int,
    payload: ExpenseCreate,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> ExpenseResponse:
    await _ensure_project(db, project_id)

    # Validate category
    cat = await db.get(BudgetCategory, payload.category_id)
    if cat is None or not cat.is_active:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid category")

    # Validate budget item if provided
    if payload.budget_item_id is not None:
        bi = await db.get(BudgetItem, payload.budget_item_id)
        if bi is None or bi.project_id != project_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid budget item")

    now = datetime.now(timezone.utc)
    expense = Expense(
        project_id=project_id,
        category_id=payload.category_id,
        budget_item_id=payload.budget_item_id,
        description=payload.description,
        amount=payload.amount,
        expense_date=payload.expense_date,
        vendor=payload.vendor,
        invoice_number=payload.invoice_number,
        notes=payload.notes,
        status=ExpenseStatus.PAID,
        paid_at=now,
        created_by=user.id,
    )
    db.add(expense)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Constraint violation")

    return _expense_to_response(await _load_expense(db, expense.id))


@router.put(
    "/expenses/{expense_id}",
    response_model=ExpenseResponse,
    summary="Update an expense",
)
async def update_expense(
    expense_id: int,
    payload: ExpenseUpdate,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> ExpenseResponse:
    expense = await _load_expense(db, expense_id)

    update_data = payload.model_dump(exclude_unset=True)

    if "category_id" in update_data:
        cat = await db.get(BudgetCategory, update_data["category_id"])
        if cat is None or not cat.is_active:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid category")

    if "budget_item_id" in update_data and update_data["budget_item_id"] is not None:
        bi = await db.get(BudgetItem, update_data["budget_item_id"])
        if bi is None or bi.project_id != expense.project_id:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid budget item")

    for field, value in update_data.items():
        setattr(expense, field, value)

    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Constraint violation")

    return _expense_to_response(await _load_expense(db, expense_id))


@router.delete(
    "/expenses/{expense_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete an expense",
    response_model=None,
)
async def delete_expense(
    expense_id: int,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
):
    expense = await db.get(Expense, expense_id)
    if expense is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Expense not found")

    await db.delete(expense)
    await db.commit()


# ---- Excel Import ----


@router.post(
    "/projects/{project_id}/expenses/import",
    response_model=ExpenseImportResult,
    summary="Bulk-import expenses from an Excel (.xlsx) file",
)
async def import_expenses_from_excel(
    project_id: int,
    db: DBSession,
    file: UploadFile = File(...),
    default_category_id: int = Form(...),
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> ExpenseImportResult:
    await _ensure_project(db, project_id)

    # Validate default category
    default_cat = await db.get(BudgetCategory, default_category_id)
    if default_cat is None or not default_cat.is_active:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid default category")

    # Validate file
    if file.content_type not in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    ):
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Only .xlsx files are accepted",
        )

    contents = await file.read()
    if len(contents) > MAX_IMPORT_FILE_SIZE:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "File exceeds 5 MB limit")

    # Parse Excel
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Could not read Excel file")

    ws = wb.active
    if ws is None:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "No active sheet in workbook")

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "File must have a header row and at least one data row",
        )

    header_row = rows[0]
    header_map = _build_header_map([str(h) if h else "" for h in header_row])

    # We need at least amount column
    if "amount" not in header_map:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"Could not find an 'amount' column. Headers found: {[str(h) for h in header_row if h]}",
        )

    # Pre-load category map for name-based lookup
    category_map = await _get_category_map(db)

    now = datetime.now(timezone.utc)
    imported = 0
    errors: list[ImportRowError] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            # Amount (required)
            raw_amount = row[header_map["amount"]] if "amount" in header_map else None
            if raw_amount is None or str(raw_amount).strip() == "":
                errors.append(ImportRowError(row=row_idx, reason="Missing amount"))
                continue
            try:
                amount = Decimal(str(raw_amount).replace(",", "").replace(" ", "").strip())
            except (InvalidOperation, ValueError):
                errors.append(ImportRowError(row=row_idx, reason=f"Invalid amount: {raw_amount}"))
                continue
            if amount <= 0:
                errors.append(ImportRowError(row=row_idx, reason=f"Amount must be positive: {raw_amount}"))
                continue

            # Expense date
            raw_date = row[header_map["expense_date"]] if "expense_date" in header_map else None
            expense_date: date
            if raw_date is None or str(raw_date).strip() == "":
                expense_date = date.today()
            elif isinstance(raw_date, datetime):
                expense_date = raw_date.date()
            elif isinstance(raw_date, date):
                expense_date = raw_date
            else:
                try:
                    # Try common formats
                    date_str = str(raw_date).strip()
                    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
                        try:
                            expense_date = datetime.strptime(date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                    else:
                        errors.append(ImportRowError(row=row_idx, reason=f"Invalid date: {raw_date}"))
                        continue
                except Exception:
                    errors.append(ImportRowError(row=row_idx, reason=f"Invalid date: {raw_date}"))
                    continue

            # Vendor (optional)
            raw_vendor = row[header_map["vendor"]] if "vendor" in header_map else None
            vendor = str(raw_vendor).strip()[:255] if raw_vendor else None

            # Invoice number (optional)
            raw_inv = row[header_map["invoice_number"]] if "invoice_number" in header_map else None
            invoice_number = str(raw_inv).strip()[:100] if raw_inv else None

            # Description (optional — fallback to vendor + invoice)
            raw_desc = row[header_map["description"]] if "description" in header_map else None
            if raw_desc and str(raw_desc).strip():
                description = str(raw_desc).strip()[:500]
            else:
                parts = [p for p in [vendor, invoice_number] if p]
                description = " — ".join(parts) if parts else f"Import row {row_idx}"

            # Category (optional — use name lookup, fallback to default_category_id)
            raw_cat = row[header_map["category"]] if "category" in header_map else None
            if raw_cat and str(raw_cat).strip():
                cat_name_lower = str(raw_cat).strip().lower()
                category_id = category_map.get(cat_name_lower, default_category_id)
            else:
                category_id = default_category_id

            expense = Expense(
                project_id=project_id,
                category_id=category_id,
                description=description,
                amount=amount,
                expense_date=expense_date,
                vendor=vendor,
                invoice_number=invoice_number,
                status=ExpenseStatus.PAID,
                paid_at=now,
                created_by=user.id,
            )
            db.add(expense)
            imported += 1

        except Exception as exc:
            errors.append(ImportRowError(row=row_idx, reason=str(exc)))

    # Bulk commit
    if imported > 0:
        try:
            await db.commit()
        except IntegrityError:
            await db.rollback()
            raise HTTPException(
                status.HTTP_400_BAD_REQUEST,
                "Database error during bulk insert — no rows were imported",
            )

    return ExpenseImportResult(
        imported_count=imported,
        skipped_count=len(errors),
        errors=errors,
    )
