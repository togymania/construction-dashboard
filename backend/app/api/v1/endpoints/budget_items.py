"""Budget item CRUD endpoints + per-project budget summary aggregation."""
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload

import io
from decimal import Decimal as _Decimal
from decimal import InvalidOperation as _InvalidOperation
from typing import Literal

from fastapi import File, Form, UploadFile
from openpyxl import load_workbook

from app.api.deps import CurrentUser, DBSession, require_roles
from app.core.config import settings
from app.services.category_service import get_or_create_category, resolve_category
from app.models.budget import BudgetItem
from app.models.budget_category import BudgetCategory
from app.models.expense import Expense, ExpenseStatus
from app.models.project import Project
from app.models.user import User, UserRole
from app.schemas.budget import (
    BudgetCategoryBreakdown,
    BudgetImportResult,
    BudgetImportRowError,
    BudgetImportRowWarning,
    BudgetItemCreate,
    BudgetItemResponse,
    BudgetItemUpdate,
    BudgetSummary,
)
from app.schemas.budget_variance import BudgetVarianceReport

router = APIRouter(tags=["Budget Items"])


# ---------- Helpers ----------

async def _ensure_project_exists(db, project_id: int) -> Project:
    project = await db.get(Project, project_id)
    if project is None or not project.is_active:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found",
        )
    return project


async def _ensure_category_exists(db, category_id: int) -> BudgetCategory:
    cat = await db.get(BudgetCategory, category_id)
    if cat is None or not cat.is_active:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Category id={category_id} does not exist or is inactive",
        )
    return cat


async def _get_budget_item_with_category(db, item_id: int) -> BudgetItem | None:
    """Fetch a budget item with its category eagerly loaded."""
    stmt = (
        select(BudgetItem)
        .options(selectinload(BudgetItem.category))
        .where(BudgetItem.id == item_id)
    )
    result = await db.execute(stmt)
    return result.scalar_one_or_none()


# ---------- BudgetItem CRUD (nested under projects) ----------

@router.get(
    "/projects/{project_id}/budget-items",
    response_model=list[BudgetItemResponse],
    summary="List budget items for a project",
)
async def list_budget_items(
    project_id: int,
    user: CurrentUser,
    db: DBSession,
) -> list[BudgetItem]:
    await _ensure_project_exists(db, project_id)

    stmt = (
        select(BudgetItem)
        .options(selectinload(BudgetItem.category))
        .join(BudgetCategory, BudgetItem.category_id == BudgetCategory.id)
        .where(BudgetItem.project_id == project_id)
        .order_by(BudgetCategory.display_order, BudgetItem.id)
    )
    result = await db.execute(stmt)
    return list(result.scalars().all())


@router.post(
    "/projects/{project_id}/budget-items",
    response_model=BudgetItemResponse,
    status_code=status.HTTP_201_CREATED,
    summary="Create a new budget item under a project",
)
async def create_budget_item(
    project_id: int,
    payload: BudgetItemCreate,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> BudgetItem:
    await _ensure_project_exists(db, project_id)

    # Resolve category: existing id OR free-text auto-create
    try:
        category = await resolve_category(
            db,
            category_id=payload.category_id,
            category_name_new=payload.category_name_new,
        )
    except ValueError as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(e))

    item = BudgetItem(
        project_id=project_id,
        category_id=category.id,
        description=payload.description,
        cost_code=payload.cost_code,
        planned_amount=payload.planned_amount,
        committed_amount=payload.committed_amount,
        notes=payload.notes,
    )
    db.add(item)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not create budget item (constraint violation)",
        )
    await db.refresh(item, attribute_names=["category"])
    return item


@router.put(
    "/projects/{project_id}/budget-items/{item_id}",
    response_model=BudgetItemResponse,
    summary="Update a budget item (scoped under project)",
)
async def update_budget_item(
    project_id: int,
    item_id: int,
    payload: BudgetItemUpdate,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> BudgetItem:
    await _ensure_project_exists(db, project_id)
    item = await _get_budget_item_with_category(db, item_id)
    if item is None or item.project_id != project_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Budget item not found",
        )

    update_data = payload.model_dump(exclude_unset=True)

    # Category change: existing id OR free-text auto-create (mutually exclusive)
    if "category_id" in update_data or "category_name_new" in update_data:
        try:
            category = await resolve_category(
                db,
                category_id=update_data.get("category_id"),
                category_name_new=update_data.get("category_name_new"),
            )
        except ValueError as e:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, str(e))
        update_data["category_id"] = category.id
        update_data.pop("category_name_new", None)  # not a real ORM field

    for field, value in update_data.items():
        setattr(item, field, value)

    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not update budget item (constraint violation)",
        )
    await db.refresh(item, attribute_names=["category"])
    return item


@router.delete(
    "/projects/{project_id}/budget-items/{item_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="Delete a budget item (scoped under project)",
)
async def delete_budget_item(
    project_id: int,
    item_id: int,
    db: DBSession,
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> None:
    await _ensure_project_exists(db, project_id)
    item = await db.get(BudgetItem, item_id)
    if item is None or item.project_id != project_id:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Budget item not found",
        )

    # Note: expenses with budget_item_id pointing here will have it set to NULL
    # (per ondelete='SET NULL' on Expense.budget_item_id)
    await db.delete(item)
    await db.commit()


# ---------- Budget Summary (aggregation) ----------

@router.get(
    "/projects/{project_id}/budget-summary",
    response_model=BudgetSummary,
    summary="Get budget summary for a project (planned vs spent, by category)",
)
async def get_budget_summary(
    project_id: int,
    user: CurrentUser,
    db: DBSession,
) -> BudgetSummary:
    """Return the project's budget summary.

    "Spent" is computed from the **ledger** (entry_type=EXPENSE) rather
    than the legacy ``expenses`` table -- the ledger is the source of
    truth on this project. We treat as project-relevant any ledger row
    whose linked contract belongs to this project, plus any unlinked
    rows (assumed project-wide costs).
    """
    from app.models.ledger_entry import LedgerEntry, LedgerEntryType
    from app.models.subcontractor import SubcontractorContract

    project = await _ensure_project_exists(db, project_id)

    # ---- Total planned (sum of all budget items) ----
    total_planned_stmt = select(
        func.coalesce(func.sum(BudgetItem.planned_amount), 0)
    ).where(BudgetItem.project_id == project_id)
    total_planned: Decimal = (await db.execute(total_planned_stmt)).scalar_one()

    # ---- Ledger-side filters: project's contracts OR unlinked entries ----
    ledger_project_filter = (SubcontractorContract.id.is_(None)) | (
        SubcontractorContract.project_id == project_id
    )

    # ---- Total spent + record count (from ledger expense entries) ----
    ledger_total_stmt = (
        select(
            func.coalesce(func.sum(LedgerEntry.amount), 0),
            func.count(LedgerEntry.id),
        )
        .outerjoin(
            SubcontractorContract,
            SubcontractorContract.id == LedgerEntry.contract_id,
        )
        .where(
            LedgerEntry.entry_type == LedgerEntryType.EXPENSE,
            ledger_project_filter,
        )
    )
    total_spent_row = (await db.execute(ledger_total_stmt)).first()
    total_spent: Decimal = Decimal(total_spent_row[0])
    expense_records_count: int = int(total_spent_row[1])

    # ---- Pending kept zero for now: ledger has no "pending" concept;
    #      pending payments live on SubcontractorPayment, modelled
    #      separately. Future: roll those in here too. ----
    total_pending: Decimal = Decimal("0")

    remaining = total_planned - total_spent
    utilization_pct = (
        float(total_spent / total_planned * 100) if total_planned > 0 else 0.0
    )

    # ---- Planned per category (still from budget_items) ----
    planned_per_cat_stmt = (
        select(
            BudgetItem.category_id,
            func.coalesce(func.sum(BudgetItem.planned_amount), 0).label("planned"),
        )
        .where(BudgetItem.project_id == project_id)
        .group_by(BudgetItem.category_id)
    )
    planned_rows = (await db.execute(planned_per_cat_stmt)).all()
    planned_by_cat: dict[int, Decimal] = {row[0]: row[1] for row in planned_rows}

    # ---- Spent per category: aggregate ledger by budget_code, then map
    #      cost_code -> category_id via this project's budget_items. ----
    ledger_per_code_stmt = (
        select(
            LedgerEntry.budget_code,
            func.coalesce(func.sum(LedgerEntry.amount), 0).label("total"),
        )
        .outerjoin(
            SubcontractorContract,
            SubcontractorContract.id == LedgerEntry.contract_id,
        )
        .where(
            LedgerEntry.entry_type == LedgerEntryType.EXPENSE,
            LedgerEntry.budget_code.is_not(None),
            ledger_project_filter,
        )
        .group_by(LedgerEntry.budget_code)
    )
    ledger_per_code_rows = (await db.execute(ledger_per_code_stmt)).all()
    ledger_by_code: dict[str, Decimal] = {
        (row[0] or "").strip().lower(): Decimal(row[1])
        for row in ledger_per_code_rows
        if row[0]
    }

    code_to_cat_stmt = select(BudgetItem.cost_code, BudgetItem.category_id).where(
        BudgetItem.project_id == project_id,
        BudgetItem.cost_code.is_not(None),
    )
    code_rows = (await db.execute(code_to_cat_stmt)).all()
    cost_code_to_cat: dict[str, int] = {
        (row[0] or "").strip().lower(): row[1]
        for row in code_rows
        if row[0]
    }

    # Kullanıcı bir harcamaya item-level cost_code ("3", "29"...) yerine
    # category-level slug ("bina", "yollar"...) atayabiliyor.
    # İkinci eşleme: category slug → category_id.
    slug_rows = (await db.execute(select(BudgetCategory.slug, BudgetCategory.id))).all()
    slug_to_cat: dict[str, int] = {
        (row[0] or "").strip().lower(): row[1]
        for row in slug_rows
        if row[0]
    }

    spent_by_cat: dict[int, Decimal] = {}
    for code, amt in ledger_by_code.items():
        cat_id = cost_code_to_cat.get(code) or slug_to_cat.get(code)
        if cat_id is None:
            continue
        spent_by_cat[cat_id] = spent_by_cat.get(cat_id, Decimal("0")) + amt

    relevant_cat_ids = set(planned_by_cat.keys()) | set(spent_by_cat.keys())
    if relevant_cat_ids:
        cat_stmt = (
            select(BudgetCategory)
            .where(BudgetCategory.id.in_(relevant_cat_ids))
            .order_by(BudgetCategory.display_order, BudgetCategory.id)
        )
        cats = list((await db.execute(cat_stmt)).scalars().all())
    else:
        cats = []

    by_category: list[BudgetCategoryBreakdown] = []
    for cat in cats:
        planned = planned_by_cat.get(cat.id, Decimal("0"))
        spent = spent_by_cat.get(cat.id, Decimal("0"))
        cat_remaining = planned - spent
        cat_util = float(spent / planned * 100) if planned > 0 else 0.0
        by_category.append(
            BudgetCategoryBreakdown(
                category_id=cat.id,
                category_name=cat.name,
                category_slug=cat.slug,
                planned_amount=planned,
                spent_amount=spent,
                remaining_amount=cat_remaining,
                utilization_pct=cat_util,
            )
        )

    return BudgetSummary(
        project_id=project_id,
        project_budget_rub=project.budget_rub,
        total_planned=total_planned,
        total_spent=total_spent,
        total_pending=total_pending,
        remaining=remaining,
        utilization_pct=utilization_pct,
        by_category=by_category,
        expense_records_count=expense_records_count,
    )


# ---------- Budget Item Excel Import ----------

# Header aliases (TR / EN / RU). All compared case-insensitively after .lower().
_BUDGET_COLUMN_ALIASES: dict[str, list[str]] = {
    "category": [
        "category", "kategori",
        "категория",  # Cyrillic
        "budget category", "bütçe kategorisi",
    ],
    "cost_code": [
        "code", "cost code", "wbs", "wbs code", "item code",
        "kod", "kalem kodu", "iş kalemi kodu",
        "код", "шифр",
    ],
    "description": [
        "item", "item name", "description", "desc",
        "kalem", "kalem adı", "açıklama",
        "наименование", "описание",
    ],
    "amount": [
        "amount", "budget", "total", "planned", "planned amount",
        "tutar", "bütçe", "miktar", "planlanan", "planlanan tutar",
        "сумма", "бюджет", "плановая",
    ],
    "committed_amount": [
        "committed", "committed amount", "commitment", "po", "po amount",
        "taahhüt", "taahhüt edilen", "kontrat tutarı",
        "обязательство", "законтрактовано",
    ],
    "notes": [
        "notes", "note",
        "not", "açıklama 2",
        "комментарий", "примечание",
    ],
}


def _build_budget_header_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        normalised = str(raw).strip().lower()
        for field, aliases in _BUDGET_COLUMN_ALIASES.items():
            if field.startswith("_"):  # ignored fields like cost code
                continue
            if normalised in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


@router.post(
    "/projects/{project_id}/budget-items/import",
    response_model=BudgetImportResult,
    summary="Bulk-import budget items from an Excel (.xlsx) file",
)
async def import_budget_items_from_excel(
    project_id: int,
    db: DBSession,
    file: UploadFile = File(...),
    overwrite_mode: Literal["append", "replace"] = Form("append"),
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> BudgetImportResult:
    """
    Bulk-create budget items from an Excel sheet.

    Modes:
      * append  : insert rows, skip duplicates
                  ((project_id, category_id, lower(description)) match)
      * replace : DELETE every existing budget item for the project first,
                  then insert all rows from the Excel.
                  WARNING: destructive. Caller must confirm in UI.

    Categories:
      * If the row's "Category" cell matches an existing budget category
        (case-insensitive, trim/normalised), that category is used.
      * Otherwise a new category is auto-created (is_system=False).
        Each newly-created category produces a single warning, not one per row.
    """
    await _ensure_project_exists(db, project_id)

    if overwrite_mode not in ("append", "replace"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="overwrite_mode must be 'append' or 'replace'",
        )

    # File extension check
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted",
        )

    max_bytes = settings.MAX_IMPORT_FILE_SIZE_MB * 1024 * 1024
    contents = await file.read()
    if len(contents) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File exceeds {settings.MAX_IMPORT_FILE_SIZE_MB} MB limit",
        )

    # Parse Excel
    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not read Excel file",
        )

    ws = wb.active
    if ws is None:
        wb.close()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No active sheet in workbook",
        )

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must have a header row and at least one data row",
        )

    header_row = rows[0]
    header_map = _build_budget_header_map([str(h) if h else "" for h in header_row])

    # Required columns
    if "amount" not in header_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Could not find an 'amount' / 'tutar' / 'сумма' column. "
                f"Headers found: {[str(h) for h in header_row if h]}"
            ),
        )
    if "description" not in header_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "Could not find a description column "
                "(item / kalem / описание / etc)."
            ),
        )

    # Replace mode: wipe existing items first.
    deleted_count = 0
    if overwrite_mode == "replace":
        existing = (await db.execute(
            select(BudgetItem).where(BudgetItem.project_id == project_id)
        )).scalars().all()
        deleted_count = len(existing)
        for item in existing:
            await db.delete(item)
        await db.flush()

    # Pre-load known category IDs (to announce auto-created ones once)
    known_category_ids: set[int] = {
        cat_id for cat_id in (
            await db.execute(select(BudgetCategory.id))
        ).scalars().all()
    }

    # Pre-load existing (category_id, lower(description)) pairs for dup detection.
    # In replace mode we just deleted everything, so this is empty.
    existing_pairs: set[tuple[int, str]] = set()
    if overwrite_mode == "append":
        rows_check = (await db.execute(
            select(BudgetItem.category_id, BudgetItem.description).where(
                BudgetItem.project_id == project_id
            )
        )).all()
        existing_pairs = {(c, (d or "").lower()) for c, d in rows_check}

    # Track in-file duplicates too
    seen_in_file: set[tuple[int, str]] = set()

    imported = 0
    errors: list[BudgetImportRowError] = []
    warnings: list[BudgetImportRowWarning] = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            # ---- Description (required) ----
            raw_desc = row[header_map["description"]]
            description = str(raw_desc).strip() if raw_desc else ""
            if not description:
                errors.append(BudgetImportRowError(
                    row=row_idx, reason="Missing description"
                ))
                continue
            description = description[:500]

            # ---- Amount (required) ----
            raw_amount = row[header_map["amount"]]
            if raw_amount is None or str(raw_amount).strip() == "":
                errors.append(BudgetImportRowError(
                    row=row_idx, reason="Missing amount"
                ))
                continue
            try:
                amount = _Decimal(
                    str(raw_amount).replace(",", "").replace(" ", "").strip()
                )
            except (_InvalidOperation, ValueError):
                errors.append(BudgetImportRowError(
                    row=row_idx, reason=f"Invalid amount: {raw_amount}"
                ))
                continue
            if amount < 0:
                errors.append(BudgetImportRowError(
                    row=row_idx, reason=f"Amount cannot be negative: {raw_amount}"
                ))
                continue

            # ---- Category (optional - auto-create) ----
            cat_id: int
            if "category" in header_map:
                raw_cat = row[header_map["category"]]
                if raw_cat and str(raw_cat).strip():
                    try:
                        cat = await get_or_create_category(db, str(raw_cat).strip())
                    except ValueError:
                        errors.append(BudgetImportRowError(
                            row=row_idx, reason="Invalid category name"
                        ))
                        continue
                    cat_id = cat.id
                    if cat.id not in known_category_ids:
                        warnings.append(BudgetImportRowWarning(
                            row=row_idx,
                            reason=f"Auto-created new category '{cat.name}'",
                        ))
                        known_category_ids.add(cat.id)
                else:
                    errors.append(BudgetImportRowError(
                        row=row_idx, reason="Category cell is empty"
                    ))
                    continue
            else:
                errors.append(BudgetImportRowError(
                    row=row_idx,
                    reason="No category column in file (Category / Kategori / Категория required)",
                ))
                continue

            # ---- Notes (optional) ----
            notes_val: str | None = None
            if "notes" in header_map:
                raw_notes = row[header_map["notes"]]
                if raw_notes and str(raw_notes).strip():
                    notes_val = str(raw_notes).strip()

            # ---- Cost code (optional, but used for planned-vs-actual) ----
            cost_code_val: str | None = None
            if "cost_code" in header_map:
                raw_code = row[header_map["cost_code"]]
                if raw_code is not None and str(raw_code).strip():
                    cost_code_val = str(raw_code).strip()[:50]

            # ---- Committed amount (optional) ----
            committed_val: _Decimal = _Decimal("0")
            if "committed_amount" in header_map:
                raw_committed = row[header_map["committed_amount"]]
                if raw_committed is not None and str(raw_committed).strip() != "":
                    try:
                        committed_val = _Decimal(
                            str(raw_committed).replace(",", "").replace(" ", "").strip()
                        )
                        if committed_val < 0:
                            committed_val = _Decimal("0")
                            warnings.append(BudgetImportRowWarning(
                                row=row_idx,
                                reason="Negative committed amount → treated as 0",
                            ))
                    except (_InvalidOperation, ValueError):
                        warnings.append(BudgetImportRowWarning(
                            row=row_idx,
                            reason=f"Invalid committed amount '{raw_committed}' → 0",
                        ))

            # ---- Duplicate detection ----
            dup_key = (cat_id, description.lower())
            if dup_key in existing_pairs:
                warnings.append(BudgetImportRowWarning(
                    row=row_idx,
                    reason=f"Skipped: duplicate of existing item ({description})",
                ))
                continue
            if dup_key in seen_in_file:
                warnings.append(BudgetImportRowWarning(
                    row=row_idx,
                    reason=f"Skipped: duplicate within file ({description})",
                ))
                continue
            seen_in_file.add(dup_key)

            # ---- Insert ----
            bi = BudgetItem(
                project_id=project_id,
                category_id=cat_id,
                description=description,
                cost_code=cost_code_val,
                planned_amount=amount,
                committed_amount=committed_val,
                notes=notes_val,
            )
            db.add(bi)
            try:
                await db.flush()
            except IntegrityError as ie:
                await db.rollback()
                errors.append(BudgetImportRowError(
                    row=row_idx,
                    reason=f"DB error: {str(ie.orig)[:200]}",
                ))
                continue

            imported += 1

        except Exception as exc:
            errors.append(BudgetImportRowError(row=row_idx, reason=str(exc)))

    # Final commit
    if imported > 0 or deleted_count > 0:
        try:
            await db.commit()
        except IntegrityError:
            await db.rollback()
            return BudgetImportResult(
                imported_count=0,
                skipped_count=len(errors) + len(warnings) + imported,
                deleted_count=0,
                errors=errors + [BudgetImportRowError(
                    row=0, reason="Final commit failed - no rows imported"
                )],
                warnings=warnings,
            )

    return BudgetImportResult(
        imported_count=imported,
        skipped_count=len(errors) + len(warnings),
        deleted_count=deleted_count,
        errors=errors,
        warnings=warnings,
    )


# ---------- Planned vs Actual variance (Faz 3) ----------


@router.get(
    "/projects/{project_id}/budget/variance",
    response_model=BudgetVarianceReport,
    summary="Planned vs actual report for every budget item on a project",
)
async def get_variance_report(
    project_id: int,
    user: CurrentUser,
    db: DBSession,
) -> BudgetVarianceReport:
    """Return the project-wide planned vs actual breakdown.

    Actuals are pulled from PAID expenses linked to the budget item, plus
    EXPENSE-type ledger entries whose `budget_code` matches the item's
    `cost_code` and whose contract belongs to this project (or is
    project-wide / unlinked).
    """
    from app.services.budget_variance import build_variance_report

    await _ensure_project_exists(db, project_id)
    return await build_variance_report(db, project_id)


# ---------- ÇMI-format Monart-only import (Faz 2 — Monart spesifik) ----------


@router.post(
    "/projects/{project_id}/budget-items/import-cmi",
    response_model=BudgetImportResult,
    summary="Import Monart's lines from a ÇMI-format master budget workbook",
)
async def import_cmi_monart(
    project_id: int,
    db: DBSession,
    file: UploadFile = File(...),
    sheet_name: str = Form("ЦМИ"),
    responsible_filter: str = Form("Монарт"),
    overwrite_mode: Literal["append", "replace"] = Form("append"),
    user: User = Depends(require_roles(UserRole.ADMIN, UserRole.PROJECT_MANAGER)),
) -> BudgetImportResult:
    """Parse the ÇMI master sheet and import only the rows whose
    "Bütçe sorumlusu" (column P) matches ``responsible_filter``.

    Monart's 15 top-level work packages get auto-categorized via
    ``monart_budget_parser.category_for(cost_code)`` (Bina, Yollar, Altyapı,
    Haberleşme, Isıtma, Elektrik, Peyzaj, Aydınlatma, Diğer İnşaat).

    Detail rows (cost code empty in the source) are NOT imported as separate
    budget items — they are folded into the parent's ``notes`` field as a
    bulleted breakdown so we don't double-count.
    """
    from app.services.monart_budget_parser import (
        parse as parse_cmi,
        category_for as cmi_category_for,
        render_detail_note,
    )

    await _ensure_project_exists(db, project_id)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are accepted",
        )

    max_bytes = settings.MAX_LEDGER_IMPORT_MB * 1024 * 1024  # ÇMI workbook is multi-MB
    contents = await file.read()
    if len(contents) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"File exceeds {settings.MAX_LEDGER_IMPORT_MB} MB limit",
        )

    try:
        report = parse_cmi(
            contents,
            sheet_name=sheet_name,
            responsible_filter=responsible_filter,
        )
    except ValueError as e:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(e))

    # Replace mode: drop existing items first
    deleted_count = 0
    if overwrite_mode == "replace":
        existing = (await db.execute(
            select(BudgetItem).where(BudgetItem.project_id == project_id)
        )).scalars().all()
        deleted_count = len(existing)
        for it in existing:
            await db.delete(it)
        await db.flush()

    # Pre-load known categories so we know which auto-creates to warn about
    known_category_ids: set[int] = {
        cid for cid in (
            await db.execute(select(BudgetCategory.id))
        ).scalars().all()
    }

    # Pre-load existing (project_id, lower(cost_code)) pairs for dedup in append mode
    existing_codes: set[str] = set()
    if overwrite_mode == "append":
        rows_check = (await db.execute(
            select(BudgetItem.cost_code).where(
                BudgetItem.project_id == project_id,
                BudgetItem.cost_code.is_not(None),
            )
        )).all()
        existing_codes = {(c or "").strip().lower() for (c,) in rows_check}

    imported = 0
    errors: list[BudgetImportRowError] = []
    warnings: list[BudgetImportRowWarning] = []

    for parsed in report.items:
        try:
            # Resolve category (auto-create when missing)
            cat_name = cmi_category_for(parsed.cost_code)
            try:
                cat = await get_or_create_category(db, cat_name)
            except ValueError:
                errors.append(BudgetImportRowError(
                    row=parsed.source_row, reason=f"Invalid category '{cat_name}'"
                ))
                continue

            if cat.id not in known_category_ids:
                warnings.append(BudgetImportRowWarning(
                    row=parsed.source_row,
                    reason=f"Auto-created category '{cat.name}'",
                ))
                known_category_ids.add(cat.id)

            # Idempotency: skip if cost_code already imported for this project
            code_key = (parsed.cost_code or "").strip().lower()
            if code_key in existing_codes:
                warnings.append(BudgetImportRowWarning(
                    row=parsed.source_row,
                    reason=f"Skipped duplicate cost code '{parsed.cost_code}'",
                ))
                continue
            existing_codes.add(code_key)

            # Build notes from detail rows
            notes_val = render_detail_note(parsed) or None

            bi = BudgetItem(
                project_id=project_id,
                category_id=cat.id,
                description=parsed.description[:500],
                cost_code=parsed.cost_code[:50] if parsed.cost_code else None,
                planned_amount=parsed.planned_amount,
                committed_amount=Decimal("0"),  # not in ÇMI source; user fills later
                notes=notes_val,
            )
            db.add(bi)
            try:
                await db.flush()
            except IntegrityError as ie:
                await db.rollback()
                errors.append(BudgetImportRowError(
                    row=parsed.source_row,
                    reason=f"DB error: {str(ie.orig)[:200]}",
                ))
                continue

            imported += 1
        except Exception as exc:
            errors.append(BudgetImportRowError(
                row=parsed.source_row, reason=str(exc)
            ))

    if imported > 0 or deleted_count > 0:
        try:
            await db.commit()
        except IntegrityError:
            await db.rollback()
            return BudgetImportResult(
                imported_count=0,
                skipped_count=len(errors) + len(warnings) + imported,
                deleted_count=0,
                errors=errors + [BudgetImportRowError(
                    row=0, reason="Final commit failed",
                )],
                warnings=warnings,
            )

    return BudgetImportResult(
        imported_count=imported,
        skipped_count=len(errors) + len(warnings),
        deleted_count=deleted_count,
        errors=errors,
        warnings=warnings,
    )
