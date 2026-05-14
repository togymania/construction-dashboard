"""Parse the OZET sheet from a 'Harcama Takip' Excel file.

Şirketin standart aylık para akışı özeti şu yapıdadır:

    R1: B = başlık (örn. "HIPODROM - OZET")
        C = as_of_date (Excel tarihi)
    R2: C = obje adı header (HIPODROM / MOSKOVA HIPODROM)
    R3: B="ISVEREN TAHSILATLARI"     C=<rakam>
    R4: B="FIRMA ODEMELERI"          C=<rakam>
    R5: B="UCRET GIDERLERI"          C=<rakam>
    R6: B="VERGI ODEMELERI"          C=<rakam>
    R7: B="Gelir Vergisi"            C=<rakam>    (alt kırılım)
    R8: B="KDV"                      C=<rakam>    (alt kırılım)
    R9: B="FAIZ GELIRLERI"           C=<rakam>
    R10: B="BANKA GIDERLERI"         C=<rakam>
    R11: B="DIGER GELIR-GIDERLER"    C=<rakam>
    R12: B="TOPLAM"                  C=<rakam>

Şirket etiketi (Monotek vs Monart) dosya adından heuristic ile
çıkarılır — Excel'in içinde net bir company id alanı yok.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

import openpyxl


# Etiket → ORM field eşlemesi.
_LABEL_MAP: dict[str, str] = {
    "ISVEREN TAHSILATLARI":  "isveren_tahsilatlari",
    "FIRMA ODEMELERI":       "firma_odemeleri",
    "UCRET GIDERLERI":       "ucret_giderleri",
    "VERGI ODEMELERI":       "vergi_odemeleri",
    "GELIR VERGISI":         "gelir_vergisi",
    "KDV":                   "kdv",
    "FAIZ GELIRLERI":        "faiz_gelirleri",
    "BANKA GIDERLERI":       "banka_giderleri",
    "DIGER GELIR-GIDERLER":  "diger_gelir_giderler",
    "TOPLAM":                "toplam",
}


class OzetParseError(Exception):
    """OZET sayfası bulunamadı veya parse edilemedi."""


def normalize_label(s: str) -> str:
    """Eşleştirme için string'i normalize et — Türkçe ı/İ duyarlılığı vs."""
    if not s:
        return ""
    # Türkçe karakterleri ASCII'ye çevir (basit)
    t = (
        s.replace("İ", "I").replace("ı", "i")
        .replace("Ş", "S").replace("ş", "s")
        .replace("Ğ", "G").replace("ğ", "g")
        .replace("Ü", "U").replace("ü", "u")
        .replace("Ö", "O").replace("ö", "o")
        .replace("Ç", "C").replace("ç", "c")
    )
    return t.strip().upper()


def detect_company_label(filename: str) -> str:
    """Dosya adından şirket etiketi tahmini.

    "Harcama Takip-HIPODROM-Monart.xlsx" → "Monart"
    "Harcama Takip-HIPODROM-Monotek.xlsx" → "Monotek"
    """
    if not filename:
        return "Unknown"
    n = filename.lower()
    if "monart" in n:
        return "Monart"
    if "monotek" in n:
        return "Monotek"
    return "Unknown"


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal(0)
    try:
        # Excel float'larında bilimsel notasyon olabiliyor → str + Decimal
        return Decimal(str(v))
    except Exception:  # noqa: BLE001
        return Decimal(0)


def _to_date(v: Any) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_ozet(raw_bytes: bytes) -> dict:
    """Excel byte içeriğinden OZET değerlerini çıkar.

    Dönüş örneği::

        {
            "as_of_date": date(2026, 3, 31),
            "isveren_tahsilatlari": Decimal("4374915305.56"),
            "firma_odemeleri":      Decimal("-2554791100.30"),
            ...
            "toplam":               Decimal("1062578480.71"),
        }

    OZET sheet yoksa veya beklenen labellar bulunamazsa OzetParseError
    fırlatır.
    """
    # read_only=True streams the workbook instead of loading every sheet
    # into memory — kritik çünkü Render free tier 512 MB ile sınırlı ve
    # Harcama Takip dosyaları 10+ MB olabiliyor.
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(raw_bytes), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001
        raise OzetParseError(f"Excel açılamadı: {exc}") from exc

    try:
        # OZET sheet adının exact case'i bilinmez, normalize ederek bul
        target = None
        for name in wb.sheetnames:
            if name.strip().upper() == "OZET":
                target = name
                break
        if target is None:
            raise OzetParseError(
                f"'OZET' sayfası bulunamadı. Mevcut sayfalar: {wb.sheetnames}"
            )

        ws = wb[target]

        # read_only mode'da random cell access yok, sıralı iter_rows lazım.
        # OZET sayfası küçük (~12 satır) — sadece ilk 30 satıra bak.
        result: dict = {"as_of_date": None}
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=30, max_col=3, values_only=True),
            start=1,
        ):
            # row: tuple of (col A, col B, col C)
            if row_idx == 1:
                # R1 C → tarih
                result["as_of_date"] = _to_date(row[2] if len(row) > 2 else None)
                continue
            label_cell = row[1] if len(row) > 1 else None
            value_cell = row[2] if len(row) > 2 else None
            if label_cell is None:
                continue
            normalized = normalize_label(str(label_cell))
            if normalized in _LABEL_MAP:
                field = _LABEL_MAP[normalized]
                result[field] = _to_decimal(value_cell)
    finally:
        wb.close()

    # En azından "toplam" bulunmuş olmalı; aksi takdirde format yanlış.
    if "toplam" not in result:
        raise OzetParseError(
            "OZET sayfasında beklenen kalemler bulunamadı (TOPLAM eksik). "
            "Şablon değişmiş olabilir."
        )

    # Boş kalanları sıfırla
    for f in _LABEL_MAP.values():
        result.setdefault(f, Decimal(0))

    # as_of_date hâlâ None ise bugünü kullan
    if result["as_of_date"] is None:
        result["as_of_date"] = date.today()

    return result
