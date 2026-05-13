"""Tender → TDF Excel (КП Форма) export.

Şirket içinde standart TDF (Teklif Değerlendirme Formu / Коммерческое
Предложение Форма) şablonunun birebir kopyasını üretir.

Layout — analyze_tdf.py ile çıkarılan gerçek 'r02' şablonundan:

    Row 1   : kurumsal banner
    Row 3-4 : (boş — şablon görsel/logo bandı bırakıyor)
    Row 5   : Bidder ad satırı, her bidder için 2 sütun merge edilir
              (Цена + Стоимость sütun çifti üstüne)
    Row 6   : (boş spacer)
    Row 7   : Тема + tender başlığı (A:G arası merged); her bidder'ın
              kolonunda BİDDER TOPLAMI üstte gösterilir
    Row 8   : Sütun başlıkları
              [Наименование | Изображение из АГР | План | Тип пирога |
               Ед. изм. | Кол-во]  +  per-bidder: [Цена монтажа за ед. |
               Общая стоимость]
    Row 9+  : Line items
    Row N+1 : spacer
    Row N+3 : Контактное лицо, ФИО, телефон, e-mail:  →  contact info
    Row N+6 : Комментарии  (her bidder kolon başlığı olarak)
    Row N+7 : В стоимость входит:    → her bidder için
    Row N+8 : В стоимость НЕ входит: → her bidder için
    Row N+12: Условия оплаты:        → her bidder için
    Row N+13: Сроки выполнения работ: → her bidder için
    Row N+14: Объект:                → tender.object_name

SUPERSEDED bid'ler dahil edilmez. Variant'lar (aynı firma + farklı
variant_label) ayrı bidder olarak yan yana iki sütun çifti alır.
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.tender import (
    Bid,
    BidLineItem,
    BidPriceType,
    BidStatus,
    Tender,
    TenderLineItem,
)


# ---------------------------------------------------------------------------
# Stil paleti — kurumsal mavi (Monotekstroy)
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style="thin", color="BCC8D6"),
    right=Side(style="thin", color="BCC8D6"),
    top=Side(style="thin", color="BCC8D6"),
    bottom=Side(style="thin", color="BCC8D6"),
)

FILL_BIDDER = PatternFill("solid", fgColor="143C73")    # bidder adı bandı
FILL_HEADER = PatternFill("solid", fgColor="1FA3DA")    # sütun başlık
FILL_TOTAL = PatternFill("solid", fgColor="E8F1FB")     # totals row
FILL_LABEL = PatternFill("solid", fgColor="F1F5F9")     # footer label
FILL_ALT = PatternFill("solid", fgColor="FAFCFE")       # zebra
FILL_PACKAGE = PatternFill("solid", fgColor="E2ECF7")   # package row

FONT_BANNER = Font(name="Calibri", size=14, bold=True, color="143C73")
FONT_BIDDER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_TOPLAM = Font(name="Calibri", size=12, bold=True, color="143C73")
FONT_THEME = Font(name="Calibri", size=12, bold=True, color="143C73")
FONT_LABEL = Font(name="Calibri", size=11, bold=True, color="143C73")
FONT_BODY = Font(name="Calibri", size=11, color="1E293B")
FONT_BODY_BOLD = Font(name="Calibri", size=11, bold=True, color="143C73")
FONT_NEGOTIABLE = Font(name="Calibri", size=10, italic=True, color="C2410C")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

# Sabit sol sütunlar
FIXED_COLS = [
    ("№",                  5),
    ("Наименование",       40),
    ("Изображение из АГР", 18),
    ("План",               10),
    ("Тип пирога",         12),
    ("Ед. изм.",           9),
    ("Кол-во",             10),
]
PER_BIDDER_COLS = [
    ("Цена монтажа\nза ед., руб. с НДС", 16),
    ("Общая стоимость,\nруб. с НДС",     18),
]
NUM_FIXED = len(FIXED_COLS)
NUM_PER_BIDDER = len(PER_BIDDER_COLS)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_tdf_workbook(tender: Tender) -> bytes:
    """Tender'ı TDF formatlı .xlsx olarak üret, byte stream döndür."""
    wb = Workbook()
    ws = wb.active
    sheet_title = (tender.title or "КП Форма")[:31].replace("/", "-").replace("\\", "-")
    ws.title = sheet_title or "КП Форма"

    bidders: list[Bid] = sorted(
        [b for b in tender.bids if b.status != BidStatus.SUPERSEDED],
        key=lambda b: (b.company_name, b.variant_label or ""),
    )
    line_items: list[TenderLineItem] = sorted(
        tender.line_items, key=lambda li: li.order_num
    )

    _write_banner(ws)
    _write_bidder_band(ws, bidders)
    _write_topic_and_totals(ws, tender, bidders)
    _write_column_headers(ws, bidders)
    last_row = _write_line_items(ws, line_items, bidders)
    _write_footer(ws, tender, bidders, last_row + 2)
    _set_widths(ws, bidders)
    _set_print_options(ws, bidders)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Section 1: kurumsal banner (Row 1-2)
# ---------------------------------------------------------------------------


def _write_banner(ws) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
    c = ws.cell(row=1, column=1, value="MONOTEKSTROY · КОММЕРЧЕСКОЕ ПРЕДЛОЖЕНИЕ")
    c.font = FONT_BANNER
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[1].height = 22


# ---------------------------------------------------------------------------
# Section 2: bidder isim bandı (Row 5)
# ---------------------------------------------------------------------------


def _bidder_col_start(bidder_idx: int) -> int:
    """Bidder #idx'in sol sütununun 1-tabanlı index'i."""
    return NUM_FIXED + 1 + bidder_idx * NUM_PER_BIDDER


def _write_bidder_band(ws, bidders: Sequence[Bid]) -> None:
    if not bidders:
        return
    row = 5
    for i, b in enumerate(bidders):
        col = _bidder_col_start(i)
        label = b.company_name
        if b.variant_label:
            label += f"\n{b.variant_label}"
        # Merge bidder's two columns
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + NUM_PER_BIDDER - 1,
        )
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = FILL_BIDDER
        cell.font = FONT_BIDDER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 32


# ---------------------------------------------------------------------------
# Section 3: Тема + bidder totals (Row 7)
# ---------------------------------------------------------------------------


def _write_topic_and_totals(ws, tender: Tender, bidders: Sequence[Bid]) -> None:
    row = 7
    # "Тема" hücresi (A7)
    label = ws.cell(row=row, column=1, value="Тема")
    label.font = FONT_LABEL
    label.alignment = ALIGN_LEFT
    label.fill = FILL_LABEL
    label.border = THIN_BORDER

    # Tender başlığı (B7 ile başlayıp G7'ye kadar merged)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NUM_FIXED)
    title_cell = ws.cell(row=row, column=2, value=tender.title or "")
    title_cell.font = FONT_THEME
    title_cell.alignment = ALIGN_LEFT
    title_cell.fill = FILL_LABEL
    title_cell.border = THIN_BORDER

    # Her bidder için toplam (üstte göze çarpsın)
    for i, b in enumerate(bidders):
        col = _bidder_col_start(i)
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + NUM_PER_BIDDER - 1,
        )
        total = Decimal(b.total_amount or 0)
        cell = ws.cell(row=row, column=col, value=float(total) if total > 0 else None)
        cell.font = FONT_TOPLAM
        cell.fill = FILL_TOTAL
        cell.alignment = ALIGN_RIGHT
        cell.border = THIN_BORDER
        cell.number_format = '#,##0.00 " ₽";-#,##0.00 " ₽";""'
    ws.row_dimensions[row].height = 26


# ---------------------------------------------------------------------------
# Section 4: sütun başlıkları (Row 8)
# ---------------------------------------------------------------------------


def _write_column_headers(ws, bidders: Sequence[Bid]) -> None:
    row = 8
    # Sabit sütun başlıkları
    for ci, (label, _w) in enumerate(FIXED_COLS, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
    # Her bidder için iki başlık
    for i, _b in enumerate(bidders):
        col = _bidder_col_start(i)
        for k, (label, _w) in enumerate(PER_BIDDER_COLS):
            c = ws.cell(row=row, column=col + k, value=label)
            c.fill = FILL_HEADER
            c.font = FONT_HEADER
            c.alignment = ALIGN_CENTER
            c.border = THIN_BORDER
    ws.row_dimensions[row].height = 38


# ---------------------------------------------------------------------------
# Section 5: line items (Row 9+)
# ---------------------------------------------------------------------------


def _write_line_items(
    ws,
    line_items: Sequence[TenderLineItem],
    bidders: Sequence[Bid],
) -> int:
    start_row = 9
    if not line_items:
        return start_row - 1
    for idx, li in enumerate(line_items):
        r = start_row + idx
        is_package = li.line_type == "package"
        is_work = li.line_type == "work"

        # Zebra / paket / normal arka plan
        if is_package:
            bg = FILL_PACKAGE
        elif idx % 2 == 1:
            bg = FILL_ALT
        else:
            bg = None

        body_font = FONT_BODY_BOLD if is_package else FONT_BODY

        # # / Описание / Изображение / План / Тип / Ед. / Кол-во
        # parent_id varsa açıklamayı iki boşlukla indentle (görsel hiyerarşi)
        indent = "    " if li.parent_id else ""
        cells = [
            (1, li.display_label or li.order_num),
            (2, indent + (li.description or "")),
            (3, ""),                              # Изображение (görsel için boş)
            (4, ""),                              # План (görsel için boş)
            (5, _line_type_label(li)),            # Тип пирога / sub-type hint
            (6, li.unit or ""),
            (7, float(li.quantity or 0) if li.quantity else None),
        ]
        for ci, val in cells:
            c = ws.cell(row=r, column=ci, value=val if val != "" else None)
            c.font = body_font
            c.alignment = ALIGN_LEFT if ci == 2 else ALIGN_CENTER
            c.border = THIN_BORDER
            if bg is not None:
                c.fill = bg
            if ci == 7 and isinstance(val, (int, float)):
                c.number_format = '#,##0.00;-#,##0.00;""'

        # Her bidder için Цена + Стоимость
        qty = Decimal(li.quantity or 0)
        for i, b in enumerate(bidders):
            col = _bidder_col_start(i)
            cell_data = _find_bid_line(b, li.id)
            if cell_data is None:
                # boş hücreler
                _empty_currency(ws, r, col, bg)
                _empty_currency(ws, r, col + 1, bg)
                continue
            if cell_data.price_type != BidPriceType.FIXED:
                # "Договорная", "не включена", ...
                text = cell_data.raw_text_price or cell_data.price_type.value
                _put_negotiable(ws, r, col, text, bg)
                _put_negotiable(ws, r, col + 1, text, bg)
                continue
            unit_price = Decimal(cell_data.unit_price_total or 0)
            total = qty * unit_price
            _put_money(ws, r, col, unit_price if unit_price else None, bg)
            _put_money(ws, r, col + 1, total if (unit_price and qty) else None, bg, bold=True)

    last_row = start_row + len(line_items) - 1
    return last_row


def _line_type_label(li: TenderLineItem) -> str:
    """Görsel olarak 'Тип пирога' kolonuna katkı — line_type hint'i."""
    if li.line_type == "work":
        return "Работы"
    if li.line_type == "material":
        return "Материал"
    if li.line_type == "package":
        return ""
    return ""


def _find_bid_line(bid: Bid, line_id: int) -> BidLineItem | None:
    for bl in bid.line_items:
        if bl.tender_line_item_id == line_id:
            return bl
    return None


def _put_money(ws, r, c, value, fill, *, bold=False):
    cell = ws.cell(row=r, column=c, value=float(value) if value is not None else None)
    cell.number_format = '#,##0.00 " ₽";-#,##0.00 " ₽";""'
    cell.font = (
        Font(name="Calibri", size=11, bold=True, color="143C73")
        if bold else FONT_BODY
    )
    cell.alignment = ALIGN_RIGHT
    cell.border = THIN_BORDER
    if fill is not None:
        cell.fill = fill
    return cell


def _empty_currency(ws, r, c, fill):
    cell = ws.cell(row=r, column=c, value=None)
    cell.alignment = ALIGN_RIGHT
    cell.border = THIN_BORDER
    if fill is not None:
        cell.fill = fill
    return cell


def _put_negotiable(ws, r, c, text, fill):
    cell = ws.cell(row=r, column=c, value=text)
    cell.font = FONT_NEGOTIABLE
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    if fill is not None:
        cell.fill = fill
    return cell


# ---------------------------------------------------------------------------
# Section 6: footer
# ---------------------------------------------------------------------------


def _write_footer(ws, tender: Tender, bidders: Sequence[Bid], start_row: int) -> None:
    """Her label tek bir satır, sağında her bidder için 2 sütun merged."""
    rows: list[tuple[str, callable]] = [
        ("Контактное лицо, ФИО, телефон, e-mail:", _contact_block),
        ("Комментарии",                            lambda b: b.notes or ""),
        ("В стоимость входит:",                    lambda b: b.included_in_price or ""),
        ("В стоимость НЕ входит:",                 lambda b: b.not_included_in_price or ""),
        ("Условия оплаты:",                        lambda b: b.payment_terms or ""),
        ("Сроки выполнения работ:",                lambda b: _delivery_text(b)),
    ]
    cur_row = start_row
    for label, getter in rows:
        # Label tarafı (A:G merge)
        ws.merge_cells(
            start_row=cur_row, start_column=1,
            end_row=cur_row, end_column=NUM_FIXED,
        )
        lc = ws.cell(row=cur_row, column=1, value=label)
        lc.font = FONT_LABEL
        lc.fill = FILL_LABEL
        lc.alignment = ALIGN_LEFT
        lc.border = THIN_BORDER

        # Her bidder kolon çifti
        for i, b in enumerate(bidders):
            col = _bidder_col_start(i)
            ws.merge_cells(
                start_row=cur_row, start_column=col,
                end_row=cur_row, end_column=col + NUM_PER_BIDDER - 1,
            )
            value = getter(b)
            vc = ws.cell(row=cur_row, column=col, value=value)
            vc.font = FONT_BODY
            vc.alignment = ALIGN_LEFT
            vc.border = THIN_BORDER

        ws.row_dimensions[cur_row].height = 45 if label.startswith(("В стоимость", "Условия")) else 28
        cur_row += 1

    # Объект (en altta tek satır)
    cur_row += 1
    ws.merge_cells(
        start_row=cur_row, start_column=1,
        end_row=cur_row, end_column=NUM_FIXED,
    )
    obj_label = ws.cell(row=cur_row, column=1, value="Объект:")
    obj_label.font = FONT_LABEL
    obj_label.fill = FILL_LABEL
    obj_label.alignment = ALIGN_LEFT
    obj_label.border = THIN_BORDER

    # Sağ tarafa tender.object_name (tüm bidder kolonları boyunca merge)
    last_col = _bidder_col_start(len(bidders) - 1) + NUM_PER_BIDDER - 1 if bidders else NUM_FIXED + 1
    if last_col > NUM_FIXED:
        ws.merge_cells(
            start_row=cur_row, start_column=NUM_FIXED + 1,
            end_row=cur_row, end_column=last_col,
        )
        oc = ws.cell(row=cur_row, column=NUM_FIXED + 1, value=tender.object_name or "")
        oc.font = FONT_BODY_BOLD
        oc.alignment = ALIGN_LEFT
        oc.border = THIN_BORDER


def _contact_block(bid: Bid) -> str:
    parts = []
    if bid.contact_name:
        parts.append(bid.contact_name)
    if bid.contact_phone:
        parts.append(bid.contact_phone)
    if bid.contact_email:
        parts.append(bid.contact_email)
    return ", ".join(parts)


def _delivery_text(bid: Bid) -> str:
    if bid.delivery_days is not None:
        return f"{bid.delivery_days} дней"
    return ""


# ---------------------------------------------------------------------------
# Section 7: sütun genişlikleri + print options
# ---------------------------------------------------------------------------


def _set_widths(ws, bidders: Sequence[Bid]) -> None:
    for ci, (_label, width) in enumerate(FIXED_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width
    for i, _b in enumerate(bidders):
        col = _bidder_col_start(i)
        for k, (_label, width) in enumerate(PER_BIDDER_COLS):
            ws.column_dimensions[get_column_letter(col + k)].width = width


def _set_print_options(ws, bidders: Sequence[Bid]) -> None:
    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    # Donmuş başlık: ilk 8 satır + 2 sol sütun (Наименование + #)
    ws.freeze_panes = "C9"
