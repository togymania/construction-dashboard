"""Parser for HIPODROM-style Gelir-Gider Excel sheets.

Reads the `Gelir-Gider` sheet of the workbook and extracts six columns:
    B (idx 1): Tarih           → entry_date
    C (idx 2): Aciklamalar     → description
    E (idx 4): Firma/Sahis     → company_name
    F (idx 5): Kod             → kod
    G (idx 6): Hesap           → account
    J (idx 9): Miktar (signed) → amount + entry_type

Header is at row 3, data starts row 4. Sign of Miktar determines direction:
positive → INCOME, negative → EXPENSE. Zero / blank Miktar rows are skipped.
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from app.models.ledger_entry import LedgerEntryType


SHEET_NAME = "Gelir-Gider"
# Header position is auto-detected (HIPODROM uses row 3, MOSKOVA HIPODROM uses
# row 2). We scan the first few rows of column B for the literal "Tarih"
# (case-insensitive) and treat that as the header row.
HEADER_SEARCH_RANGE = 6  # scan rows 1..6 for the header
HEADER_KEYWORD = "tarih"

# Column indexes (0-indexed for tuple-style row access)
COL_DATE = 1          # B
COL_DESC = 2          # C
COL_COMPANY = 4       # E
COL_KOD = 5           # F
COL_ACCOUNT = 6       # G
COL_AMOUNT = 9        # J
# Hidden columns used for dedup only — not displayed in the UI.
# The source procurement system assigns a unique payment_id per actual payment;
# two partial payments against the same invoice will share invoice/date/amount
# but have distinct payment_ids, so this is the most reliable dedup key.
COL_PAYMENT_ID = 36   # AK ("ID платежа")


@dataclass
class ParsedLedgerRow:
    """A single parsed ledger row, ready for preview/persist."""

    source_row: int
    entry_date: date
    description: str | None
    company_name: str | None
    kod: str | None
    account: str | None
    amount: Decimal
    entry_type: LedgerEntryType
    dedup_hash: str


@dataclass
class RowError:
    source_row: int
    reason: str


@dataclass
class ParseResult:
    rows: list[ParsedLedgerRow] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)
    skipped_blank: int = 0          # all-None data row
    skipped_zero_amount: int = 0    # Miktar is 0 / blank


# ---- Helpers -----------------------------------------------------------------


def _to_date(raw) -> date | None:
    """Coerce various date encodings to a Python date, or None on blank."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = str(raw).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_decimal(raw) -> Decimal | None:
    """Coerce a numeric cell to Decimal. Returns None on blank, raises on garbage."""
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))
    text = str(raw).strip()
    if not text:
        return None
    # Strip thousand separators / spaces; tolerate comma decimals
    cleaned = text.replace(" ", "").replace("\xa0", "")
    # If both . and , present, assume , is thousand sep (RU/TR varies)
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    return Decimal(cleaned)


def _clean_str(raw, max_len: int | None = None) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    return text[:max_len] if max_len else text


def compute_dedup_hash(
    *,
    payment_id: str | None,
    entry_date: date,
    entry_type: LedgerEntryType,
    amount: Decimal,
    company_name: str | None,
    description: str | None,
) -> str:
    """Stable hash for duplicate detection across re-imports of the same Excel.

    Strategy:
      - If `payment_id` is present (column AK / "ID платежа"), combine it
        with date + entry_type + amount. The source system reuses the same
        payment_id for related transactions (e.g. a payment and its later
        refund), so payment_id alone collides those legitimate pairs;
        adding date + type + amount disambiguates them while still catching
        true duplicates on re-import.
      - Otherwise (legacy / non-system rows with no payment_id), fall back
        to a composite hash of date | type | amount | company | description.

    Yields hex SHA-256.
    """
    if payment_id:
        payload = "|".join(
            [
                "pid",
                payment_id.strip(),
                entry_date.isoformat(),
                entry_type.value,
                f"{amount:.2f}",
            ]
        )
    else:
        # Use the FULL description — many ledgers tag each row with a unique
        # incoming-reference number (e.g. "Вх.номер: 48594") at the end of the
        # description that disambiguates otherwise-identical template rows
        # (e.g. per-employee government fees with the same template prefix).
        desc_part = (description or "").strip().lower()
        comp_part = (company_name or "").strip().lower()
        payload = "|".join(
            [
                "noid",
                entry_date.isoformat(),
                entry_type.value,
                f"{amount:.2f}",
                comp_part,
                desc_part,
            ]
        )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


# ---- Public API --------------------------------------------------------------


def parse_gelir_gider(file_bytes: bytes) -> ParseResult:
    """Parse the Gelir-Gider sheet of the given .xlsx bytes.

    Returns a ParseResult with parsed rows, per-row errors, and skip counts.
    Does NOT touch the database.
    """
    result = ParseResult()

    try:
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        result.errors.append(RowError(source_row=0, reason=f"Could not open workbook: {exc}"))
        return result

    if SHEET_NAME not in wb.sheetnames:
        result.errors.append(
            RowError(
                source_row=0,
                reason=f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}",
            )
        )
        wb.close()
        return result

    ws = wb[SHEET_NAME]

    # Auto-detect header row: scan first few rows for "Tarih" in column B
    header_row_idx: int | None = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_RANGE, values_only=True), start=1):
        cell_b = row[COL_DATE] if len(row) > COL_DATE else None
        if cell_b and isinstance(cell_b, str) and cell_b.strip().lower() == HEADER_KEYWORD:
            header_row_idx = r_idx
            break
    if header_row_idx is None:
        result.errors.append(
            RowError(
                source_row=0,
                reason=f"Could not find header row (no 'Tarih' label in column B within first {HEADER_SEARCH_RANGE} rows)",
            )
        )
        wb.close()
        return result

    data_start_row = header_row_idx + 1

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx < data_start_row:
            continue

        # Treat row as blank if all of date/desc/company/amount are None
        key_cells = (
            row[COL_DATE] if len(row) > COL_DATE else None,
            row[COL_DESC] if len(row) > COL_DESC else None,
            row[COL_COMPANY] if len(row) > COL_COMPANY else None,
            row[COL_AMOUNT] if len(row) > COL_AMOUNT else None,
        )
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in key_cells):
            result.skipped_blank += 1
            continue

        # ---- Amount (drives type) ----
        raw_amount = row[COL_AMOUNT] if len(row) > COL_AMOUNT else None
        try:
            amount_signed = _to_decimal(raw_amount)
        except (InvalidOperation, ValueError):
            result.errors.append(
                RowError(source_row=row_idx, reason=f"Invalid Miktar: {raw_amount!r}")
            )
            continue

        if amount_signed is None or amount_signed == 0:
            result.skipped_zero_amount += 1
            continue

        entry_type = LedgerEntryType.INCOME if amount_signed > 0 else LedgerEntryType.EXPENSE
        amount = abs(amount_signed)

        # ---- Date ----
        raw_date = row[COL_DATE] if len(row) > COL_DATE else None
        entry_date = _to_date(raw_date)
        if entry_date is None:
            result.errors.append(
                RowError(source_row=row_idx, reason=f"Invalid/missing Tarih: {raw_date!r}")
            )
            continue

        # ---- Optional strings ----
        description = _clean_str(row[COL_DESC] if len(row) > COL_DESC else None)
        company_name = _clean_str(row[COL_COMPANY] if len(row) > COL_COMPANY else None, max_len=500)
        kod = _clean_str(row[COL_KOD] if len(row) > COL_KOD else None, max_len=50)
        account = _clean_str(row[COL_ACCOUNT] if len(row) > COL_ACCOUNT else None, max_len=100)

        # Hidden — used for dedup only
        payment_id = _clean_str(row[COL_PAYMENT_ID] if len(row) > COL_PAYMENT_ID else None)

        dedup = compute_dedup_hash(
            payment_id=payment_id,
            entry_date=entry_date,
            entry_type=entry_type,
            amount=amount,
            company_name=company_name,
            description=description,
        )

        result.rows.append(
            ParsedLedgerRow(
                source_row=row_idx,
                entry_date=entry_date,
                description=description,
                company_name=company_name,
                kod=kod,
                account=account,
                amount=amount,
                entry_type=entry_type,
                dedup_hash=dedup,
            )
        )

    wb.close()
    return result


def deduplicate_within_file(rows: list[ParsedLedgerRow]) -> tuple[list[ParsedLedgerRow], int]:
    """Drop rows whose dedup_hash already appeared earlier in the same parse.

    Returns (deduped_rows, count_dropped).
    """
    seen: set[str] = set()
    out: list[ParsedLedgerRow] = []
    dropped = 0
    for r in rows:
        if r.dedup_hash in seen:
            dropped += 1
            continue
        seen.add(r.dedup_hash)
        out.append(r)
    return out, dropped
