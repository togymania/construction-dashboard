"""Excel parser for daily workforce snapshots in the Monotekstroy/Monart cover-page format.

The cover page sheet has three sections (A: PRODUCTIVE, B: UNPRODUCTIVE, C: SUBCONT)
each followed by a TOTAL row. We walk rows, switch state when we see a section
marker, and emit (category, position_name, general_staff, absent, leave_sick, present)
tuples for each data row.

Reference shape (header row anchors):
    "РАБОЧИЕ" or "PRODUCTIVE" -> Section A starts (DIRECT)
    "PRODUCTIVE LABOUR" -> Section A ends, totals on this row
    "UNPRODUCTIVE PERSONEL" -> Section B starts (INDIRECT)
    "UNPRODUCTIVE LABOUR" -> Section B ends, totals
    "SUBCONT. PRODUCTIVE" -> Section C starts (SUBCONTRACTOR)
    "TOTAL SUBCONTRACTOR" -> Section C ends, totals
    "GRAND TOTAL" -> overall validation row
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Iterable

from openpyxl import load_workbook


# ---------- Result types ----------

@dataclass
class ParsedRow:
    """A single position+counts row extracted from the cover page."""
    category: str  # "direct" | "indirect" | "subcontractor"
    position_name: str
    general_staff: int
    absent: int
    leave_sick: int
    present: int


@dataclass
class ParsedTotals:
    """A section/grand total row, used for validation."""
    label: str  # "PRODUCTIVE LABOUR" | "UNPRODUCTIVE LABOUR" | "TOTAL SUBCONTRACTOR" | "GRAND TOTAL"
    general_staff: int
    absent: int
    leave_sick: int
    present: int


@dataclass
class ParserResult:
    snapshot_date: date | None
    project_label: str | None  # company/project text from header row, e.g. "Монотекстрой / ИППОДРОМ"
    company_label: str | None = None  # normalized to "Monotekstroy" or "Monart" (None if not detected)
    rows: list[ParsedRow] = field(default_factory=list)
    section_totals: list[ParsedTotals] = field(default_factory=list)
    grand_total: ParsedTotals | None = None
    parse_warnings: list[str] = field(default_factory=list)


# ---------- Constants ----------

_SECTION_A_MARKERS = ("РАБОЧИЕ", "PRODUCTIVE PERSONEL")  # cyrillic + latin variants seen
_SECTION_A_END = "PRODUCTIVE LABOUR"
_SECTION_B_MARKERS = ("ИТР", "UNPRODUCTIVE PERSONEL")
_SECTION_B_END = "UNPRODUCTIVE LABOUR"
_SECTION_C_MARKERS = ("SUBCONT.",)
_SECTION_C_END = "TOTAL SUBCONTRACTOR"
_GRAND_TOTAL_MARKER = "GRAND TOTAL"

# Header phrases in the column-headers row (we use this row to find the 4 count columns)
_HEADER_GENERAL = ("GENERAL STAFF", "ОБЩИЙ КАДРОВ")
_HEADER_ABSENT = ("OUT OF WORKING", "НЕ ВЫШЛИ")
_HEADER_LEAVE = ("LEAVE", "БАЛЬНИЧНЫЙ", "HOLIDAY", "SICK")
_HEADER_TOTAL = ("TOTAL", "ОБЩЕЕ")

# Possible cover sheet names (some files have trailing space etc)
_COVER_SHEET_CANDIDATES = ("cover page", "coverpage", "cover_page", "cover")


# ---------- Helpers ----------

def _str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _upper_str(v) -> str:
    return _str(v).upper()


def _is_int_like(v) -> bool:
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return not (isinstance(v, float) and (v != v))  # filter NaN
    s = _str(v)
    if not s:
        return False
    return bool(re.fullmatch(r"-?\d+(\.\d+)?", s))


def _to_int(v, default: int = 0) -> int:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return default
        return int(v)
    s = _str(v)
    if not s:
        return default
    try:
        return int(float(s))
    except ValueError:
        return default


def _excel_serial_to_date(serial: int | float) -> date | None:
    """Convert Excel serial date (days since 1899-12-30) to a date."""
    try:
        # openpyxl already returns datetime.date when cell type is date.
        # When numeric, convert manually.
        n = int(serial)
        if n < 1 or n > 80000:
            return None
        # Excel epoch hack: serial 60 was the fake leap-year-1900 day; openpyxl handles this
        from datetime import timedelta
        return date(1899, 12, 30) + timedelta(days=n)
    except (TypeError, ValueError):
        return None


def _detect_cover_sheet(wb) -> str:
    """Find the cover-page sheet by name (case-insensitive, ignore spaces)."""
    for name in wb.sheetnames:
        normalized = name.strip().lower().replace(" ", "")
        if any(c.replace(" ", "") in normalized for c in _COVER_SHEET_CANDIDATES):
            return name
    # Fallback: first sheet that mentions "PRODUCTIVE" anywhere
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True, max_row=15):
            joined = " ".join(_upper_str(c) for c in row if c is not None)
            if "PRODUCTIVE" in joined and "PERSONEL" in joined:
                return name
    raise ValueError("Could not find a Cover page sheet")


def _find_count_columns(rows: list[tuple], section_a_idx: int) -> tuple[int, int, int, int]:
    """Find indices of the 4 count columns (general, absent, leave, total) by scanning
    rows just before section A for matching header phrases.
    Returns 0-based column indices.
    """
    # Look in the 5 rows preceding section A (which is where the column headers sit)
    search_window = rows[max(0, section_a_idx - 5):section_a_idx + 1]
    for row in search_window:
        upper = [_upper_str(c) for c in row]
        general_idx = absent_idx = leave_idx = total_idx = None
        for i, cell in enumerate(upper):
            if any(m in cell for m in _HEADER_GENERAL):
                general_idx = i
            elif any(m in cell for m in _HEADER_ABSENT):
                absent_idx = i
            elif any(m in cell for m in _HEADER_LEAVE):
                leave_idx = i
            elif any(m in cell for m in _HEADER_TOTAL) and total_idx is None:
                # 'TOTAL' header - take last occurrence to avoid 'OBSCHEE' false-match
                total_idx = i
        if general_idx is not None and total_idx is not None:
            # If absent/leave missing, fall back to general+1, general+2
            if absent_idx is None:
                absent_idx = general_idx + 1
            if leave_idx is None:
                leave_idx = general_idx + 2
            return general_idx, absent_idx, leave_idx, total_idx
    # Fallback: hard-coded based on observed templates (D=3, E=4, F=5, G=6 zero-based)
    return 3, 4, 5, 6


def _extract_position_name(row: tuple, count_col_start: int) -> str:
    """Extract the position label - the longest non-empty cell BEFORE the count columns."""
    candidates = [_str(c) for c in row[:count_col_start] if _str(c)]
    if not candidates:
        return ""
    # The position label is usually the rightmost non-empty cell before counts
    return candidates[-1]


# ---------- Main parser ----------

def parse_workforce_excel(contents: bytes) -> ParserResult:
    """Parse an .xlsx daily-puantaj cover page into a ParserResult.

    Tolerant to small layout shifts; surfaces problems via parse_warnings.
    """
    result = ParserResult(snapshot_date=None, project_label=None)

    try:
        wb = load_workbook(filename=io.BytesIO(contents), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel: {e}") from e

    try:
        sheet_name = _detect_cover_sheet(wb)
    except ValueError as e:
        wb.close()
        raise

    ws = wb[sheet_name]
    rows: list[tuple] = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 5:
        raise ValueError("Cover page is too short to be a valid puantaj")

    # 1. Header parse: project label (first non-empty cell in first 3 rows) + date (numeric or date cell)
    for r in rows[:5]:
        for cell in r:
            if isinstance(cell, datetime):
                result.snapshot_date = cell.date()
                break
            if isinstance(cell, date):
                result.snapshot_date = cell
                break
        if result.snapshot_date is not None:
            break
    # If no native date, scan for an integer in the 30000-80000 range (Excel serial dates)
    if result.snapshot_date is None:
        for r in rows[:5]:
            for cell in r:
                if isinstance(cell, (int, float)) and 30000 <= cell <= 80000:
                    result.snapshot_date = _excel_serial_to_date(cell)
                    if result.snapshot_date is not None:
                        break
            if result.snapshot_date is not None:
                break

    # Project label: first long-ish text cell in row 0-1
    for r in rows[:2]:
        for cell in r:
            s = _str(cell)
            if len(s) >= 5 and not _is_int_like(cell):
                result.project_label = s
                break
        if result.project_label:
            break

    # Detect company_label from the project label (or from rows[0:3] more broadly).
    # Files have either Cyrillic "МОНОТЕКСТРОЙ" / "МОНАРТ" or Latin equivalents.
    company_search_text = ""
    for r in rows[:3]:
        for cell in r:
            company_search_text += " " + _upper_str(cell)

    if "МОНОТЕКСТРОЙ" in company_search_text or "MONOTEKSTROY" in company_search_text or "MOНOTEKSTROY" in company_search_text:
        result.company_label = "Monotekstroy"
    elif "МОНАРТ" in company_search_text or "MONART" in company_search_text:
        result.company_label = "Monart"
    else:
        result.parse_warnings.append(
            f"Could not detect company (Monotekstroy/Monart) from header. project_label={result.project_label!r}"
        )

    # 2. Locate section A start
    section_a_idx = None
    for i, r in enumerate(rows):
        joined_upper = " ".join(_upper_str(c) for c in r if c is not None)
        if any(m in joined_upper for m in _SECTION_A_MARKERS) and "PERSONEL" in joined_upper:
            section_a_idx = i
            break
    if section_a_idx is None:
        raise ValueError("Could not locate Section A (PRODUCTIVE PERSONEL) - is this a valid puantaj?")

    # 3. Find count columns by scanning headers near section A
    gen_col, abs_col, leave_col, tot_col = _find_count_columns(rows, section_a_idx)
    count_col_start = min(gen_col, abs_col, leave_col, tot_col)

    # 4. Walk rows, switch state on section markers and totals
    current_category: str | None = None  # None | "direct" | "indirect" | "subcontractor"
    in_section = False  # True between section_start and section_end_total

    for i, r in enumerate(rows):
        joined_upper = " ".join(_upper_str(c) for c in r if c is not None)

        # State transitions
        # NOTE: word-boundary checks - "PRODUCTIVE" is a substring of "UNPRODUCTIVE",
        # so we explicitly negate UNPRODUCTIVE when checking for productive markers.
        if any(m in joined_upper for m in _SECTION_A_MARKERS) and "PERSONEL" in joined_upper and "UNPRODUCTIVE" not in joined_upper:
            current_category = "direct"
            in_section = True
            continue
        if _SECTION_A_END in joined_upper and "UNPRODUCTIVE" not in joined_upper and current_category == "direct":
            # Capture totals on this row
            general = _to_int(r[gen_col]) if gen_col < len(r) else 0
            absent_v = _to_int(r[abs_col]) if abs_col < len(r) else 0
            leave_v = _to_int(r[leave_col]) if leave_col < len(r) else 0
            total_v = _to_int(r[tot_col]) if tot_col < len(r) else 0
            result.section_totals.append(ParsedTotals(
                label=_SECTION_A_END,
                general_staff=general,
                absent=absent_v,
                leave_sick=leave_v,
                present=total_v,
            ))
            in_section = False
            current_category = None
            continue
        if any(m in joined_upper for m in _SECTION_B_MARKERS) and "PERSONEL" in joined_upper:
            current_category = "indirect"
            in_section = True
            continue
        if _SECTION_B_END in joined_upper and current_category == "indirect":
            general = _to_int(r[gen_col]) if gen_col < len(r) else 0
            absent_v = _to_int(r[abs_col]) if abs_col < len(r) else 0
            leave_v = _to_int(r[leave_col]) if leave_col < len(r) else 0
            total_v = _to_int(r[tot_col]) if tot_col < len(r) else 0
            result.section_totals.append(ParsedTotals(
                label=_SECTION_B_END,
                general_staff=general,
                absent=absent_v,
                leave_sick=leave_v,
                present=total_v,
            ))
            in_section = False
            current_category = None
            continue
        if any(m in joined_upper for m in _SECTION_C_MARKERS) and "PERSONNEL" in joined_upper:
            current_category = "subcontractor"
            in_section = True
            continue
        if _SECTION_C_END in joined_upper and current_category == "subcontractor":
            general = _to_int(r[gen_col]) if gen_col < len(r) else 0
            absent_v = _to_int(r[abs_col]) if abs_col < len(r) else 0
            leave_v = _to_int(r[leave_col]) if leave_col < len(r) else 0
            total_v = _to_int(r[tot_col]) if tot_col < len(r) else 0
            result.section_totals.append(ParsedTotals(
                label=_SECTION_C_END,
                general_staff=general,
                absent=absent_v,
                leave_sick=leave_v,
                present=total_v,
            ))
            in_section = False
            current_category = None
            continue
        if _GRAND_TOTAL_MARKER in joined_upper:
            general = _to_int(r[gen_col]) if gen_col < len(r) else 0
            absent_v = _to_int(r[abs_col]) if abs_col < len(r) else 0
            leave_v = _to_int(r[leave_col]) if leave_col < len(r) else 0
            total_v = _to_int(r[tot_col]) if tot_col < len(r) else 0
            result.grand_total = ParsedTotals(
                label="GRAND TOTAL",
                general_staff=general,
                absent=absent_v,
                leave_sick=leave_v,
                present=total_v,
            )
            continue

        # Data rows: only when in_section
        if not in_section or current_category is None:
            continue

        # Skip header-like rows (PRODUCTIVE PERSONEL LIST, GENERAL STAFF column-header, etc)
        if any(h in joined_upper for h in ("PERSONEL LIST", "GENERAL STAFF", "OUT OF WORKING")):
            continue

        # Need a position name AND at least general_staff > 0 (or any non-zero count)
        position_name = _extract_position_name(r, count_col_start)
        if not position_name:
            continue

        general = _to_int(r[gen_col]) if gen_col < len(r) else 0
        absent_v = _to_int(r[abs_col]) if abs_col < len(r) else 0
        leave_v = _to_int(r[leave_col]) if leave_col < len(r) else 0
        total_v = _to_int(r[tot_col]) if tot_col < len(r) else 0

        # Skip empty rows (all zeros)
        if general == 0 and absent_v == 0 and leave_v == 0 and total_v == 0:
            continue

        result.rows.append(ParsedRow(
            category=current_category,
            position_name=position_name,
            general_staff=general,
            absent=absent_v,
            leave_sick=leave_v,
            present=total_v if total_v > 0 else max(0, general - absent_v - leave_v),
        ))

    if not result.rows:
        result.parse_warnings.append("No data rows extracted from cover page")

    return result
