"""Monart-only budget parser for the ÇMI sheet (Faz 2 — version 2).

Stripped-down, single-pass parser. Produces top-level work-package rows
where:
  - column A (cost code) is non-empty
  - column P (Bütçe sorumlusu) matches the responsible filter

Detail rows (cost code empty) are attached to the most recent matching
parent as informational sub-items.
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook


@dataclass
class MonartItem:
    cost_code: str
    description: str
    unit: str | None
    quantity: Decimal | None
    planned_amount: Decimal
    responsible: str
    sub_items: list[dict[str, Any]] = field(default_factory=list)
    source_row: int = 0


@dataclass
class MonartParseReport:
    items: list[MonartItem] = field(default_factory=list)
    total: Decimal = Decimal("0")
    detail_attached: int = 0
    skipped_no_amount: int = 0
    distinct_responsibles: list[str] = field(default_factory=list)


_MONART_CATEGORY: dict[str, str] = {
    "3": "Bina",
    "29": "Yollar",
    "30": "Yollar",
    "31": "Yollar",
    "32": "Yollar",
    "33": "Yollar",
    "35": "Altyapı",
    "36": "Altyapı",
    "37": "Altyapı",
    "38": "Haberleşme",
    "39": "Isıtma",
    "40": "Elektrik",
    "41": "Elektrik",
    "44": "Peyzaj",
    "45": "Aydınlatma",
}


def category_for(cost_code: str) -> str:
    return _MONART_CATEGORY.get(cost_code, "Diğer İnşaat")


def parse(
    contents: bytes,
    *,
    sheet_name: str = "ЦМИ",
    responsible_filter: str = "Монарт",
    code_col: int = 0,
    desc_col: int = 1,
    unit_col: int = 2,
    qty_col: int = 3,
    amount_col: int = 8,
    responsible_col: int = 15,
) -> MonartParseReport:
    """Parse the ÇMI sheet, return only rows whose Bütçe sorumlusu matches.

    A "top-level row" is one where column A has a value. Empty-A rows that
    follow are attached as informational sub-items to the previous parent.
    """
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    needle = responsible_filter.strip().lower()
    report = MonartParseReport()
    distinct: set[str] = set()
    last_parent: MonartItem | None = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Pad to the longest column we touch
        max_idx = max(code_col, desc_col, unit_col, qty_col, amount_col, responsible_col)
        cells = list(row) + [None] * max(0, max_idx + 1 - len(row))

        a_raw = cells[code_col]
        b_raw = cells[desc_col]
        amount = _decimal(cells[amount_col])
        responsible = cells[responsible_col]
        if responsible:
            distinct.add(str(responsible).strip())

        # Has a cost code? = is this a top-level (parent) candidate
        has_code = a_raw not in (None, "")

        if not has_code:
            # Empty-A row → attach to current parent if we have one
            description = (str(b_raw).strip() if b_raw else "")
            if last_parent and amount and description and amount > 0:
                last_parent.sub_items.append(
                    {"row": row_idx, "description": description[:300], "amount": float(amount)}
                )
                report.detail_attached += 1
            # Crucial: never create a top-level item from an empty-A row.
            continue

        # Top-level candidate. Reject anything not Monart.
        responsible_str = str(responsible or "").strip().lower()
        if not responsible or needle not in responsible_str:
            last_parent = None
            continue

        if amount is None or amount <= 0:
            report.skipped_no_amount += 1
            last_parent = None
            continue

        description = (str(b_raw).strip() if b_raw else "").strip()
        if not description:
            last_parent = None
            continue

        item = MonartItem(
            cost_code=_format_code(a_raw),
            description=description,
            unit=(str(cells[unit_col]).strip() if cells[unit_col] else None),
            quantity=_decimal(cells[qty_col]),
            planned_amount=amount,
            responsible=str(responsible).strip(),
            source_row=row_idx,
        )
        report.items.append(item)
        report.total += amount
        last_parent = item

    wb.close()
    report.distinct_responsibles = sorted(distinct)
    return report


def render_detail_note(item: MonartItem) -> str:
    if not item.sub_items:
        return ""
    lines = ["Detaylar:"]
    for s in item.sub_items:
        amt = f"{s['amount']:,.0f}" if s.get("amount") else "—"
        d = s.get("description") or "?"
        if len(d) > 200:
            d = d[:197] + "…"
        lines.append(f"• {d} — {amt} ₽")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------


def _format_code(raw: Any) -> str:
    if isinstance(raw, (int, float)) and float(raw).is_integer():
        return str(int(raw))
    return str(raw).strip()


def _decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw).replace(",", "").replace(" ", "").strip())
    except (InvalidOperation, ValueError):
        return None
