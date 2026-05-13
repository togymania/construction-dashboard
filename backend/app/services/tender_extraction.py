"""Convert an uploaded tender quotation file into a structured draft.

Supports:
* Excel (.xlsx, .xlsm) -- openpyxl reads the sheet, we flatten cells
  to a compact text grid for Claude.
* PDF (.pdf) -- pdfplumber pulls text per page, we feed Claude the
  raw text.

The output is a ``TenderExtraction`` (defined in
``app.schemas.tender``) which the frontend renders as an editable
draft. Nothing is persisted at this stage -- the user reviews, fixes
any AI mistakes, then submits the cleaned-up payload through
``POST /projects/{id}/tenders``.
"""
from __future__ import annotations

import io
import json
import re
from decimal import Decimal
from typing import Any

from app.core.config import settings
from app.models.tender import TenderLineItem
from app.schemas.tender import (
    ExtractedBid,
    ExtractedBidLine,
    ExtractedLineItem,
    TenderExtraction,
)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def extract_tender_from_file(
    *,
    raw_bytes: bytes,
    filename: str,
    lang: str = "EN",
) -> TenderExtraction:
    """Return the AI-built draft for a tender quotation file.

    On any failure (file unreadable, no API key, parse error) returns
    a near-empty TenderExtraction with the failure recorded in
    ``warnings`` so the frontend can still display *something*.
    """
    warnings: list[str] = []
    text_blob, fallback_title = _file_to_text(raw_bytes, filename, warnings)

    api_key = (settings.ANTHROPIC_API_KEY or "").strip()
    if not api_key:
        warnings.append("ANTHROPIC_API_KEY missing -- returning empty draft.")
        return TenderExtraction(
            title=fallback_title,
            source_filename=filename,
            source="rule",
            warnings=warnings,
        )

    try:
        parsed = _llm_extract(text_blob, fallback_title, api_key, lang=lang)
    except Exception as exc:  # noqa: BLE001 - bubble up as warning
        warnings.append(f"LLM extraction failed: {exc!r}")
        return TenderExtraction(
            title=fallback_title,
            source_filename=filename,
            source="rule",
            warnings=warnings,
        )

    parsed.source_filename = filename
    if warnings:
        parsed.warnings = warnings + parsed.warnings
    return parsed


# ---------------------------------------------------------------------------
# File -> text helpers
# ---------------------------------------------------------------------------


def _file_to_text(raw: bytes, filename: str, warnings: list[str]) -> tuple[str, str]:
    """Return (claude_input_text, fallback_title)."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    fallback_title = _strip_ext(filename)

    if ext in {"xlsx", "xlsm"}:
        try:
            text, t = _excel_to_text(raw)
            return text, t or fallback_title
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"Excel parse failed: {exc!r}")
            return "", fallback_title

    if ext == "pdf":
        try:
            text = _pdf_to_text(raw)
            return text, fallback_title
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"PDF parse failed: {exc!r}")
            return "", fallback_title

    warnings.append(
        f"Unsupported file extension '.{ext}'. "
        "Upload .xlsx / .xlsm / .pdf for best results."
    )
    # Last-resort: treat as UTF-8 text
    try:
        return raw.decode("utf-8", errors="ignore"), fallback_title
    except Exception:
        return "", fallback_title


def _strip_ext(name: str) -> str:
    return name.rsplit(".", 1)[0] if "." in name else name


def _excel_to_text(raw: bytes) -> tuple[str, str]:
    """Flatten the first sheet's cells into a compact tab-separated grid.

    Empty rows are kept (as blank lines) so Claude can see the layout.
    Cell values are stringified; merged-cell anchors keep their value
    (downstream blanks let the LLM know "this is the same as above").
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    if not wb.sheetnames:
        return "", ""
    ws = wb[wb.sheetnames[0]]

    rows_out: list[str] = []
    title_guess = ""
    max_col = min(ws.max_column or 0, 40)  # hard cap so Claude doesn't drown
    max_row = min(ws.max_row or 0, 400)

    for r in range(1, max_row + 1):
        cells: list[str] = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                cells.append("")
            else:
                s = str(v).strip().replace("\t", " ").replace("\n", " ")
                # Heuristic title guess: first "Тема" / "Konu" / "Subject" row
                if not title_guess and s.lower() in {"тема", "konu", "subject", "тема:", "konu:"}:
                    nxt = ws.cell(row=r, column=c + 1).value
                    if nxt:
                        title_guess = str(nxt).strip()
                cells.append(s)
        rows_out.append("\t".join(cells))

    return "\n".join(rows_out), title_guess


def _pdf_to_text(raw: bytes) -> str:
    import pdfplumber

    out: list[str] = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            out.append(f"--- Page {i} ---\n{txt}")
    return "\n\n".join(out)


# ---------------------------------------------------------------------------
# LLM extraction
# ---------------------------------------------------------------------------


def _llm_extract(
    text_blob: str,
    fallback_title: str,
    api_key: str,
    *,
    lang: str = "EN",
) -> TenderExtraction:
    import anthropic  # type: ignore

    client = anthropic.Anthropic(
        api_key=api_key, timeout=settings.LLM_TIMEOUT_SECONDS
    )
    # Keep the text within Claude's reasonable input range
    snippet = text_blob[:60_000]
    lang_name = "Turkish" if lang.upper() == "TR" else "English"
    prompt = _build_extraction_prompt(snippet, lang_name)

    msg = client.messages.create(
        model=settings.ANTHROPIC_MODEL,
        max_tokens=4000,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = msg.content[0].text if msg.content else "{}"
    parsed = json.loads(_extract_json_block(raw))

    return _shape_to_extraction(parsed, fallback_title)


def _build_extraction_prompt(text_blob: str, lang_name: str) -> str:
    """Compose the structured-extraction prompt for Claude.

    Pricing rule: per line, return ``unit_price_labor`` and
    ``unit_price_material`` ONLY when the source itself splits them.
    When the source has a single combined unit price, return only
    ``unit_price_total`` and leave the other two NULL. Never invent the
    split.

    Hierarchy: real КП Formaları often have rollup rows ("1. Asphalt
    paving") with child rows under them ("1.1 prep", "1.2 mix"). The
    AI should preserve that structure via ``parent_order_num``.

    Variants: the same company sometimes submits two proposals (Material
    A vs Material B). Each variant becomes a separate ExtractedBid with
    a distinct ``variant_label``.

    Text prices: when the source says "Договорная" / "не включена" /
    "по запросу" instead of a number, return that as ``price_type`` ≠
    "fixed" and copy the wording into ``raw_text_price``.
    """
    return (
        "You are a structured-extraction assistant for a construction "
        "company. Below is the raw text or grid of a tender quotation "
        "file (a 'КП Форма' / 'Teklif Değerlendirme Formu'). "
        "Extract the work package and every company's quotation into the "
        "JSON shape below.\n\n"
        f"Write any descriptive / narrative text in {lang_name}. Keep "
        "company names, proper nouns and units in the original script "
        "(e.g. 'ООО Стройка', 'm²').\n\n"
        "Hierarchy rules — IMPORTANT:\n"
        "  - Russian/Turkish tender forms often have HIERARCHICAL rows: "
        "an outer 'package' row (e.g. '1. Демонтаж', '2. Покраска') "
        "with child rows beneath it ('1.1', '1.2', '2.a', etc.).\n"
        "  - Give every row a unique integer `order_num` (1, 2, 3, ...) "
        "AND a free-form `display_label` that mirrors the source "
        "numbering ('1', '1.1', '2.a').\n"
        "  - Set `parent_order_num` to the `order_num` of the parent "
        "row when the row is a child. Top-level rows leave it null.\n"
        "  - Classify each row's `line_type`:\n"
        "      * 'package' → an outer rollup with children (no own price)\n"
        "      * 'work'    → a labor-only sub-line ('Работы по монтажу')\n"
        "      * 'material'→ a material/equipment sub-line\n"
        "      * 'misc'    → flat single row or anything else\n\n"
        "Pricing rules — IMPORTANT:\n"
        "  - When the source breaks a line item's unit price into "
        "labor (işçilik / трудозатраты / Работы) AND material "
        "(malzeme / материал / Материал), return BOTH unit_price_labor "
        "and unit_price_material; unit_price_total = labor + material.\n"
        "  - When the source gives only a single combined unit price, "
        "return ONLY unit_price_total and leave unit_price_labor / "
        "unit_price_material as null. NEVER guess the split.\n"
        "  - When the source says 'Договорная' / 'не включена' / 'по "
        "запросу' / 'TBD' / 'Anlaşmalı' / 'на согласование' instead of "
        "a number, set `price_type` to 'negotiable' (Договорная / "
        "Anlaşmalı), 'not_included' (не включена / не входит / hariç), "
        "or 'on_request' (по запросу / TBD); leave numeric fields at 0 "
        "and copy the original wording into `raw_text_price`. For "
        "normal numeric prices `price_type` is 'fixed' and you can "
        "omit `raw_text_price`.\n"
        "  - When a company didn't quote a line, omit that "
        "ExtractedBidLine entry entirely (don't fill zeros).\n\n"
        "Variant rule — IMPORTANT:\n"
        "  - Some bidders submit MULTIPLE proposals for the same tender "
        "(e.g. 'Sistem A: Dairy Plus' vs 'Sistem B: Terras'). Each "
        "variant becomes a SEPARATE entry in `bids` with the SAME "
        "`company_name` but a distinct `variant_label` describing what "
        "makes it different (material brand, sistem adı, etc.). When "
        "the bidder only has one offer, omit `variant_label`.\n\n"
        "VAT (НДС) rule:\n"
        "  - When the file states an НДС rate (most often 20%) capture "
        "it as `vat_rate`. If the bidder gives a total 'с НДС' / 'KDV "
        "dahil', put it in `total_with_vat`; the net 'без НДС' / 'KDV "
        "hariç' goes in `total_without_vat`. Default vat_rate is 20.\n\n"
        "TDF КП-Форма field dictionary (use these exact labels to map fields):\n"
        "  Header rows (Russian / Turkish):\n"
        "    'Тема' / 'Konu'                      → title (work package, top of sheet, after 'Кому/От кого/Дата')\n"
        "    'Объект' / 'Obje'                    → object_name (project / building)\n"
        "    'Контактное лицо, ФИО, телефон, e-mail' → contact_name + contact_phone + contact_email (parsed from one cell)\n"
        "    'В стоимость входит' / 'Fiyata dahil' → included_in_price\n"
        "    'В стоимость НЕ входит' / 'Fiyata dahil DEĞİL' → not_included_in_price\n"
        "    'Условия оплаты' / 'Ödeme koşulları' → payment_terms\n"
        "    'Сроки выполнения работ' / 'İş süresi' → delivery info (parse integer days into delivery_days if 'X рабочих дней' / 'X iş günü')\n"
        "    'Комментарии' / 'Yorum'              → bid.notes\n"
        "  Table columns:\n"
        "    '№'                                   → order_num\n"
        "    'Наименование' / 'Наименование товара/услуг' / 'Adı'  → description\n"
        "    'Ед. изм.' / 'Единица измерения' / 'Birim' → unit (m², компл., шт., услуга)\n"
        "    'Кол-во' / 'Adet' / 'Мест'            → quantity\n"
        "    'Цена за единицу' / 'Цена мат. за ед.' / 'Birim fiyat' → unit_price_total (or unit_price_material when row is material-only)\n"
        "    'Общая стоимость' / 'Стоимость' / 'Toplam' → derive, do NOT store separately; line_total = qty * unit_price_total\n"
        "  VAT cues:\n"
        "    'с НДС 20%' / 'с НДС 22%' / 'с НДС 18%' → vat_rate (parse the number)\n"
        "    'без НДС' / 'KDV hariç'              → total_without_vat\n"
        "    'с НДС' / 'KDV dahil' / 'ИТОГО с НДС' → total_with_vat (and total_amount in storage)\n\n"
        "PDF brand-variant cue (IMPORTANT for catalog-style КП):\n"
        "  - If the PDF/header carries a product brand or model name (e.g. 'АгроЦентрик Дэйри Плюс', "
        "'АгроЦентрик Террас', 'Knauf Aquapanel', 'TechnoNICOL XPS-35'), capture that brand/model into "
        "`variant_label`. The same supplier sending two PDFs with two different brand names is a classic "
        "two-variant submission — output two separate bids with identical company_name and different variant_label.\n\n"
        "File content:\n"
        "```\n"
        f"{text_blob}\n"
        "```\n\n"
        "Return ONLY this JSON object, no prose around it:\n"
        "{\n"
        '  "title": "the work package title (e.g. \\"Karot işleri\\")",\n'
        '  "object_name": "project / building name if mentioned",\n'
        '  "currency": "RUB | EUR | USD | TRY",\n'
        '  "payment_terms_expected": "if specified",\n'
        '  "delivery_terms_expected": "if specified",\n'
        '  "notes": "any other useful header notes",\n'
        '  "line_items": [\n'
        "    {\n"
        '      "order_num": 1,\n'
        '      "display_label": "1",\n'
        '      "parent_order_num": null,\n'
        '      "line_type": "package",\n'
        '      "description": "Демонтажные работы",\n'
        '      "unit": null,\n'
        '      "quantity": 0\n'
        "    },\n"
        "    {\n"
        '      "order_num": 2,\n'
        '      "display_label": "1.1",\n'
        '      "parent_order_num": 1,\n'
        '      "line_type": "work",\n'
        '      "description": "Демонтаж старого покрытия",\n'
        '      "unit": "м²",\n'
        '      "quantity": 120.5\n'
        "    }\n"
        "  ],\n"
        '  "bids": [\n'
        "    {\n"
        '      "company_name": "ООО АгроЦентрик",\n'
        '      "variant_label": "Dairy Plus",\n'
        '      "contact_name": "...",\n'
        '      "contact_phone": "...",\n'
        '      "contact_email": "...",\n'
        '      "included_in_price": "...",\n'
        '      "not_included_in_price": "...",\n'
        '      "payment_terms": "...",\n'
        '      "delivery_days": 30,\n'
        '      "vat_rate": 20,\n'
        '      "total_without_vat": null,\n'
        '      "total_with_vat": null,\n'
        '      "notes": "...",\n'
        '      "lines": [\n'
        '        { "order_num": 2, "unit_price_labor": 200, "unit_price_material": 800, "unit_price_total": 1000, "price_type": "fixed" },\n'
        '        { "order_num": 3, "unit_price_labor": null, "unit_price_material": null, "unit_price_total": 0, "price_type": "negotiable", "raw_text_price": "Договорная" }\n'
        "      ]\n"
        "    }\n"
        "  ],\n"
        '  "warnings": ["any ambiguity you want the user to verify"]\n'
        "}\n"
    )


def _extract_json_block(raw: str) -> str:
    raw = raw.strip()
    fence = re.search(r"```(?:json)?\s*(.*?)\s*```", raw, flags=re.DOTALL)
    if fence:
        raw = fence.group(1)
    first = raw.find("{")
    last = raw.rfind("}")
    if first >= 0 and last > first:
        return raw[first : last + 1]
    return raw or "{}"


def _shape_to_extraction(parsed: dict[str, Any], fallback_title: str) -> TenderExtraction:
    """Validate Claude's dict into our Pydantic shape with graceful fallbacks."""
    valid_line_types = {"package", "work", "material", "misc"}
    line_items: list[ExtractedLineItem] = []
    for li in parsed.get("line_items") or []:
        try:
            line_type = str(li.get("line_type") or "misc").lower()
            if line_type not in valid_line_types:
                line_type = "misc"
            line_items.append(
                ExtractedLineItem(
                    order_num=int(li.get("order_num") or len(line_items) + 1),
                    description=str(li.get("description") or "").strip() or "—",
                    unit=str(li.get("unit") or "").strip() or None,
                    quantity=_to_decimal(li.get("quantity")),
                    display_label=_str_or_none(li.get("display_label")),
                    parent_order_num=_to_int_opt(li.get("parent_order_num")),
                    line_type=line_type,  # type: ignore[arg-type]
                )
            )
        except Exception:
            continue

    bids: list[ExtractedBid] = []
    for b in parsed.get("bids") or []:
        try:
            lines: list[ExtractedBidLine] = _shape_bid_lines(b.get("lines") or [])
            vat_rate = _to_decimal_opt(b.get("vat_rate"))
            bids.append(
                ExtractedBid(
                    company_name=str(b.get("company_name") or "Unknown").strip(),
                    variant_label=_str_or_none(b.get("variant_label")),
                    contact_name=_str_or_none(b.get("contact_name")),
                    contact_phone=_str_or_none(b.get("contact_phone")),
                    contact_email=_str_or_none(b.get("contact_email")),
                    included_in_price=_str_or_none(b.get("included_in_price")),
                    not_included_in_price=_str_or_none(b.get("not_included_in_price")),
                    payment_terms=_str_or_none(b.get("payment_terms")),
                    delivery_days=_to_int_opt(b.get("delivery_days")),
                    vat_rate=vat_rate if vat_rate is not None else Decimal("20"),
                    total_without_vat=_to_decimal_opt(b.get("total_without_vat")),
                    total_with_vat=_to_decimal_opt(b.get("total_with_vat")),
                    notes=_str_or_none(b.get("notes")),
                    lines=lines,
                )
            )
        except Exception:
            continue

    return TenderExtraction(
        title=str(parsed.get("title") or fallback_title or "Untitled tender").strip(),
        object_name=_str_or_none(parsed.get("object_name")),
        currency=(str(parsed.get("currency") or "RUB").strip().upper() or "RUB"),
        payment_terms_expected=_str_or_none(parsed.get("payment_terms_expected")),
        delivery_terms_expected=_str_or_none(parsed.get("delivery_terms_expected")),
        notes=_str_or_none(parsed.get("notes")),
        line_items=line_items,
        bids=bids,
        source="llm",
        warnings=[str(w) for w in (parsed.get("warnings") or []) if w],
    )


def _shape_bid_lines(raw_lines: list[Any]) -> list[ExtractedBidLine]:
    """Parse a Claude-returned bid-lines array into ExtractedBidLine.

    Handles the new ``price_type`` + ``raw_text_price`` fields and
    falls back to "fixed" when omitted (the common case).
    """
    valid_pt = {"fixed", "negotiable", "not_included", "on_request"}
    out: list[ExtractedBidLine] = []
    for bl in raw_lines:
        try:
            pt_raw = str(bl.get("price_type") or "fixed").lower()
            pt = pt_raw if pt_raw in valid_pt else "fixed"
            raw_text = _str_or_none(bl.get("raw_text_price"))
            if pt != "fixed":
                # Non-fixed prices: numeric fields nulled / zeroed
                out.append(
                    ExtractedBidLine(
                        order_num=int(bl.get("order_num") or 0),
                        unit_price_labor=None,
                        unit_price_material=None,
                        unit_price_total=Decimal(0),
                        price_type=pt,  # type: ignore[arg-type]
                        raw_text_price=raw_text,
                    )
                )
                continue
            lab = _to_decimal_opt(bl.get("unit_price_labor"))
            mat = _to_decimal_opt(bl.get("unit_price_material"))
            tot = _to_decimal(bl.get("unit_price_total"))
            if (lab is not None or mat is not None) and tot == 0:
                tot = (lab or Decimal(0)) + (mat or Decimal(0))
            out.append(
                ExtractedBidLine(
                    order_num=int(bl.get("order_num") or 0),
                    unit_price_labor=lab,
                    unit_price_material=mat,
                    unit_price_total=tot,
                    price_type="fixed",
                    raw_text_price=raw_text,
                )
            )
        except Exception:
            continue
    return out


# ---------------------------------------------------------------------------
# Per-bid extraction (single company's quote against an existing tender)
# ---------------------------------------------------------------------------


def extract_bid_from_file(
    *,
    raw_bytes: bytes,
    filename: str,
    line_items: list[TenderLineItem],
    lang: str = "EN",
) -> ExtractedBid:
    """Extract a single company's bid from a quote file.

    Unlike :func:`extract_tender_from_file`, the tender already exists and
    its line items are known; the LLM's job is just to identify the
    bidder, pull contact info and quote terms, and match each priced row
    to one of the supplied ``line_items`` by description similarity. The
    matched prices come back keyed by ``order_num`` so the caller can
    look up the real ``tender_line_item.id``.
    """
    warnings: list[str] = []
    text_blob, fallback_company = _file_to_text(raw_bytes, filename, warnings)

    api_key = (settings.ANTHROPIC_API_KEY or "").strip()
    if not api_key:
        return ExtractedBid(
            company_name=fallback_company or "Unknown",
            notes="ANTHROPIC_API_KEY missing — manual entry required",
        )

    try:
        bid = _llm_extract_single_bid(
            text_blob,
            line_items,
            fallback_company,
            api_key,
            lang=lang,
        )
        return bid
    except Exception as exc:  # noqa: BLE001
        return ExtractedBid(
            company_name=fallback_company or "Unknown",
            notes=f"AI extraction failed: {exc!r}",
        )


def _llm_extract_single_bid(
    text_blob: str,
    line_items: list[TenderLineItem],
    fallback_company: str,
    api_key: str,
    *,
    lang: str = "EN",
) -> ExtractedBid:
    import anthropic  # type: ignore

    client = anthropic.Anthropic(
        api_key=api_key, timeout=settings.LLM_TIMEOUT_SECONDS
    )
    snippet = text_blob[:60_000]
    lang_name = "Turkish" if lang.upper() == "TR" else "English"

    # Present the existing line items so Claude can match against them
    items_block = "\n".join(
        f"  #{li.order_num} | {li.description} | unit: {li.unit or '-'} | qty: {li.quantity}"
        for li in line_items
    )

    prompt = (
        "You are a structured-extraction assistant for a construction "
        "company. The user already created a tender with the line items "
        "listed below. The attached file is ONE company's quotation for "
        "this tender. Your job is to:\n"
        "  1. Identify the company name and contact info.\n"
        "  2. If the bidder submitted this offer as one of several "
        "alternatives (e.g. 'Sistem A: Dairy Plus' vs 'Sistem B: "
        "Terras', or 'Вариант 1: материал X'), capture the short name "
        "of THIS variant into `variant_label`. If the offer has no "
        "alternative branding, leave variant_label null.\n"
        "  3. For each priced line in the file, match it to one of the "
        "tender line items below by description similarity and return "
        "the unit price keyed by `order_num`. If a line in the file "
        "doesn't match any tender line item, skip it (don't invent "
        "matches).\n"
        "  4. If the source separates işçilik (labor / Работы) and "
        "malzeme (material / Материал) per line, fill both "
        "unit_price_labor and unit_price_material (total = labor + "
        "material). If only a combined unit price is given, fill ONLY "
        "unit_price_total and leave labor/material as null.\n"
        "  5. When the source says 'Договорная' / 'не включена' / 'по "
        "запросу' / 'Anlaşmalı' instead of a number, set `price_type` "
        "to 'negotiable' / 'not_included' / 'on_request' accordingly, "
        "leave numeric fields at 0, and copy the wording into "
        "`raw_text_price`. Normal numeric prices use price_type "
        "'fixed' (you can omit it).\n"
        "  6. If the file states an НДС rate (typically 20%), capture "
        "it as `vat_rate`; default 20 if unstated. Capture the "
        "with-VAT and without-VAT totals separately when both are "
        "given.\n\n"
        "  7. TDF КП-Форма field dictionary — map source labels to JSON fields:\n"
        "     'Тема'→title (for context), 'Объект'→object_name, 'Контактное лицо ФИО телефон e-mail'→contact_name+phone+email,\n"
        "     'В стоимость входит'→included_in_price, 'В стоимость НЕ входит'→not_included_in_price,\n"
        "     'Условия оплаты'→payment_terms, 'Сроки выполнения работ' (e.g. '10 рабочих дней')→delivery_days,\n"
        "     'Ед. изм.'/'Единица измерения'→unit, 'Кол-во'/'Мест'→quantity,\n"
        "     'Цена за единицу'/'Цена мат. за ед.'→unit_price_total, 'Стоимость с НДС N%'→vat_rate=N (parse the number).\n"
        "  8. PDF brand-variant cue: if the header / first page carries a specific product brand or model name "
        "(e.g. 'АгроЦентрик Дэйри Плюс', 'АгроЦентрик Террас', 'Knauf Aquapanel'), capture that brand into "
        "`variant_label`. Two PDFs from the same supplier with different brand names mean two variants.\n\n"
        f"Write narrative fields in {lang_name}; keep company names "
        "and units in their original script.\n\n"
        "TENDER LINE ITEMS (match against these):\n"
        f"{items_block}\n\n"
        "BIDDER FILE CONTENT:\n"
        "```\n"
        f"{snippet}\n"
        "```\n\n"
        "Return ONLY this JSON object (no prose around it):\n"
        "{\n"
        '  "company_name": "...",\n'
        '  "variant_label": null,\n'
        '  "contact_name": "...",\n'
        '  "contact_phone": "...",\n'
        '  "contact_email": "...",\n'
        '  "included_in_price": "what the bidder said is included",\n'
        '  "not_included_in_price": "...",\n'
        '  "payment_terms": "...",\n'
        '  "delivery_days": <int|null>,\n'
        '  "vat_rate": 20,\n'
        '  "total_without_vat": null,\n'
        '  "total_with_vat": null,\n'
        '  "notes": "...",\n'
        '  "lines": [\n'
        '    { "order_num": 1, "unit_price_labor": null, "unit_price_material": null, "unit_price_total": 1500, "price_type": "fixed" },\n'
        '    { "order_num": 2, "unit_price_total": 0, "price_type": "negotiable", "raw_text_price": "Договорная" }\n'
        "  ]\n"
        "}\n"
    )
    msg = client.messages.create(
        model=settings.ANTHROPIC_MODEL,
        max_tokens=2500,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = msg.content[0].text if msg.content else "{}"
    parsed = json.loads(_extract_json_block(raw))

    valid_orders = {li.order_num for li in line_items}
    lines = [
        bl for bl in _shape_bid_lines(parsed.get("lines") or [])
        if bl.order_num in valid_orders
    ]
    vat_rate = _to_decimal_opt(parsed.get("vat_rate"))
    return ExtractedBid(
        company_name=str(parsed.get("company_name") or fallback_company or "Unknown").strip(),
        variant_label=_str_or_none(parsed.get("variant_label")),
        contact_name=_str_or_none(parsed.get("contact_name")),
        contact_phone=_str_or_none(parsed.get("contact_phone")),
        contact_email=_str_or_none(parsed.get("contact_email")),
        included_in_price=_str_or_none(parsed.get("included_in_price")),
        not_included_in_price=_str_or_none(parsed.get("not_included_in_price")),
        payment_terms=_str_or_none(parsed.get("payment_terms")),
        delivery_days=_to_int_opt(parsed.get("delivery_days")),
        vat_rate=vat_rate if vat_rate is not None else Decimal("20"),
        total_without_vat=_to_decimal_opt(parsed.get("total_without_vat")),
        total_with_vat=_to_decimal_opt(parsed.get("total_with_vat")),
        notes=_str_or_none(parsed.get("notes")),
        lines=lines,
    )


# ---------------------------------------------------------------------------
# Tiny coercion helpers
# ---------------------------------------------------------------------------


def _clean_num_str(v: Any) -> str:
    """Normalise a number-string from Russian/Turkish locale.

    Handles patterns like:
      * "5 917,00"  → "5917.00"   (space thousands + comma decimal)
      * "5\u00a0917.00" → "5917.00" (NBSP thousands)
      * "5,917.00"  → "5917.00"   (comma thousands)
      * "17 751 000" → "17751000"
      * "₽ 5 917"   → "5917"
    """
    s = str(v).strip()
    # strip common currency / unit markers and whitespace classes
    for ch in ("\u00a0", "\u202f", " ", "\t", "₽", "$", "€", "р.", "руб.", "руб", "RUB", "TRY", "EUR", "USD"):
        s = s.replace(ch, "")
    # If both ',' and '.' present, assume '.' is decimal and ',' is thousands
    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    return s


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal(0)
    try:
        return Decimal(_clean_num_str(v))
    except Exception:
        return Decimal(0)


def _to_decimal_opt(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(_clean_num_str(v))
    except Exception:
        return None


def _to_int_opt(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None


def _str_or_none(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None
