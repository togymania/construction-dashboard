"""ЦМИ-format budget Excel parser (Faz 2 — Monart Stroy spesifik).

The ÇMI sheet of the master budget workbook is a roll-up table where
column A is the cost code, column B is the description, column I is
the final RUB amount, and column P is the **bütçe sorumlusu** —
the responsible firm. We only want rows where the responsible firm
matches the user-provided filter (default: "Монарт") so we end up
with the lines our own company will execute.

Within those rows there are two flavours:

* **Top-level work packages** — column A has a cost code (e.g. "3",
  "29", "44"), column C has a unit ("1 комплекс"), column I has the
  rolled-up total. These become BudgetItem rows.
* **Detail sub-rows** — column A is empty, but the row still belongs
  to the previous responsible firm (because openpyxl carries the P
  cell value in merged regions). These are the breakdown of the parent
  package; we *don't* import them as separate items (would double-count
  vs. the parent's roll-up) but we attach them as text to the parent's
  ``notes`` field for traceability.

Output: a list of dicts ready to be passed straight to ``BudgetItem(**)``
plus a parser report (warnings, skipped rows, totals).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook


# Column indexes in the ÇMI sheet (0-based, after row[i] indexing).
COL_CODE = 0           # A — № / cost code
COL_DESC = 1           # B — Наименование работ и затрат
COL_UNIT = 2           # C — Ед. изм.
COL_QTY = 3            # D — Кол-во
COL_UNIT_PRICE = 4     # E — Ед. Цена
COL_AMOUNT = 8         # I — Стоимость (final)
COL_NOTE = 9           # J — Примечание
COL_RESPONSIBLE = 15   # P — Bütçe sorumlusu
COL_SOURCE = 16        # Q — Fiyat Kaynağı


@dataclass
class CmiParseRow:
    """One parsed row destined to become a BudgetItem."""

    cost_code: str | None
    description: str
    unit: str | None
    quantity: Decimal | None
    planned_amount: Decimal
    responsible: str
    source: str | None
    sub_items: list[dict[str, Any]] = field(default_factory=list)
    source_row: int = 0


@dataclass
class CmiParseResult:
    items: list[CmiParseRow]
    skipped_no_amount: int = 0
    skipped_no_match: int = 0
    detail_rows_attached: int = 0
    total_amount: Decimal = Decimal("0")
    distinct_responsibles_seen: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def parse_cmi_sheet(
    contents: bytes,
    *,
    sheet_name: str = "ЦМИ",
    responsible_filter: str = "Монарт",
) -> CmiParseResult:
    """Parse a ÇMI-format workbook and return rows for the given responsible firm.

    ``responsible_filter`` is matched as a case-insensitive substring against
    column P. Default ``"Монарт"`` catches both ``ООО Монарт Строй`` and
    ``ООО "Монарт Строй"`` (the file uses both spellings).
    """
    wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    result = CmiParseResult(items=[])
    needle = responsible_filter.strip().lower()
    distinct_resp: set[str] = set()

    # We track the most recent matching parent so detail rows (no cost_code)
    # can attach themselves to it.
    last_match: CmiParseRow | None = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        # Defensive: pad short rows
        cells = list(row) + [None] * max(0, COL_SOURCE + 1 - len(row))

        responsible = cells[COL_RESPONSIBLE]
        if responsible:
            distinct_resp.add(str(responsible).strip())

        # Decide if this row is a (a) top-level Monart row, (b) sub-detail
        # row of the previous Monart row, or (c) something we ignore.
        responsible_str = str(responsible or "").strip().lower()
        is_match = bool(responsible) and needle in responsible_str

        cost_code_raw = cells[COL_CODE]
        cost_code = _normalize_cost_code(cost_code_raw)
        description = (str(cells[COL_DESC]).strip() if cells[COL_DESC] else "")
        amount = _parse_decimal(cells[COL_AMOUNT])

        # ---- Detail row: any row WITHOUT a cost code is treated as a child
        #      of the most recently captured parent. We don't trust the P
        #      column here — merged regions in the source file copy the
        #      "Bütçe sorumlusu" string down across all detail rows, so it
        #      can't be used to distinguish parent vs. child. ----
        if not cost_code:
            if last_match is not None and amount and description:
                last_match.sub_items.append(
                    {
                        "row": row_idx,
                        "description": description,
                        "amount": float(amount),
                    }
                )
                result.detail_rows_attached += 1
            continue

        # ---- Top-level matching row (must have its own cost code AND
        #      a responsible-firm match) ----
        if not is_match:
            last_match = None  # break the parent chain
            continue

        if amount is None or amount <= 0:
            result.skipped_no_amount += 1
            last_match = None
            continue

        if not description:
            result.warnings.append(f"Row {row_idx}: missing description, skipped")
            last_match = None
            continue

        item = CmiParseRow(
            cost_code=cost_code,
            description=description,
            unit=str(cells[COL_UNIT]).strip() if cells[COL_UNIT] else None,
            quantity=_parse_decimal(cells[COL_QTY]),
            planned_amount=amount,
            responsible=str(responsible).strip(),
            source=(
                str(cells[COL_SOURCE]).strip() if cells[COL_SOURCE] else None
            ),
            source_row=row_idx,
        )
        result.items.append(item)
        result.total_amount += amount
        last_match = item

    wb.close()
    result.distinct_responsibles_seen = sorted(distinct_resp)
    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _normalize_cost_code(raw: Any) -> str | None:
    """Return a cleaned cost code string or None if the cell is empty."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        # Avoid "3.0" — write as int when possible
        if float(raw).is_integer():
            return str(int(raw))
        return str(raw).strip()
    s = str(raw).strip()
    return s or None


def _parse_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    try:
        return Decimal(str(raw).replace(",", "").replace(" ", "").strip())
    except (InvalidOperation, ValueError):
        return None


def render_sub_items_note(item: CmiParseRow) -> str:
    """Render the sub-items list as a human-readable note string."""
    if not item.sub_items:
        return ""
    lines = ["Detaylar:"]
    for s in item.sub_items:
        amt = f"{s['amount']:,.0f}" if s.get("amount") else "—"
        desc = s.get("description") or "?"
        # cap description length so DB column (Text) doesn't bloat
        desc = desc if len(desc) <= 200 else (desc[:197] + "…")
        lines.append(f"• {desc} — {amt} ₽")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Cost-code → category heuristic (for Monart specifically)
# ---------------------------------------------------------------------------

# Friendly category names for the 14 Monart top-level packages we expect.
# Used when auto-creating BudgetCategory rows. Unmatched codes fall back to
# "Other Construction".
_MONART_CATEGORY_BY_CODE: dict[str, str] = {
    "3": "Building",
    "29": "Roads",
    "30": "Roads",
    "31": "Roads",
    "32": "Roads",
    "33": "Roads",
    "35": "Utilities",
    "36": "Utilities",
    "37": "Utilities",
    "38": "Communications",
    "39": "Heating",
    "40": "Electrical",
    "41": "Electrical",
    "44": "Landscaping",
    "45": "Lighting",
}


def category_name_for_code(cost_code: str | None) -> str:
    """Pick a sensible category for a Monart cost code. Defaults to a catch-all."""
    if cost_code is None:
        return "Other Construction"
    return _MONART_CATEGORY_BY_CODE.get(cost_code.strip(), "Other Construction")
