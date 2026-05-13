"""Verilen TDF Excel'i hücre hücre döker — Claude'a yapıştırmak için.

Kullanım:
    cd C:\\Projects\\construction-dashboard\\presentation
    python -m pip install openpyxl
    python analyze_tdf.py "C:\\Users\\tolga.topal\\Downloads\\20260325_TDF_IP_ Монтаж плит_r02.xlsx"

Veya dosya yolu vermezsen presentation/ klasöründe "sample_tdf.xlsx"
arar. Output stdout'a yazılır — terminale tıkla, Ctrl+A → Ctrl+C ile
kopyala, sohbete yapıştır.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl yüklü değil. Kur:")
    print("  python -m pip install openpyxl")
    sys.exit(1)

HERE = Path(__file__).parent

if len(sys.argv) >= 2:
    path = Path(sys.argv[1])
else:
    candidates = list(HERE.glob("*.xlsx"))
    if not candidates:
        print("Dosya yolu ver: python analyze_tdf.py <path-to-xlsx>")
        sys.exit(1)
    path = candidates[0]

print(f"=== Dosya: {path.name} ===")
wb = openpyxl.load_workbook(path, data_only=True)
print(f"Sheets: {wb.sheetnames}")
print()

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n{'='*70}")
    print(f"SHEET: {sn}   rows={ws.max_row}   cols={ws.max_column}")
    print(f"{'='*70}")
    # Sütun harfleri başlığı
    cols = min(ws.max_column or 0, 18)
    header = "      | " + " | ".join(
        openpyxl.utils.get_column_letter(c).rjust(20) for c in range(1, cols + 1)
    )
    print(header)
    print("-" * len(header))
    for r in range(1, min(ws.max_row, 50) + 1):
        row_vals = []
        for c in range(1, cols + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                row_vals.append("".rjust(20))
            else:
                s = str(v).replace("\n", " ").replace("\t", " ").strip()
                if len(s) > 20:
                    s = s[:17] + "..."
                row_vals.append(s.rjust(20))
        line = "R" + str(r).zfill(3) + ": | " + " | ".join(row_vals)
        if any(x.strip() for x in row_vals):
            print(line)
    # merged cells
    print()
    if ws.merged_cells.ranges:
        print("Merged cell ranges:")
        for mr in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
            print(f"  {mr}")
